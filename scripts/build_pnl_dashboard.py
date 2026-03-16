import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# COLOR PALETTE (matches existing TotalGuard branding)
# ============================================================
DARK_GREEN = "1B4332"
LIGHT_GREEN_BG = "D8F3DC"
MED_GREEN = "95D5B2"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
BORDER_COLOR = "CCCCCC"
BLACK = "000000"

# Reusable styles
font_title = Font(name="Arial", size=16, bold=True, color=DARK_GREEN)
font_header = Font(name="Arial", size=11, bold=True, color=WHITE)
font_section = Font(name="Arial", size=11, bold=True, color=BLACK)
font_normal = Font(name="Arial", size=10, color=BLACK)
font_bold = Font(name="Arial", size=10, bold=True, color=BLACK)
font_italic = Font(name="Arial", size=10, italic=True, color=BLACK)
font_total_big = Font(name="Arial", size=12, bold=True, color=BLACK)

fill_header = PatternFill("solid", fgColor=DARK_GREEN)
fill_title = PatternFill("solid", fgColor=LIGHT_GREEN_BG)
fill_med_green = PatternFill("solid", fgColor=MED_GREEN)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_gray = PatternFill("solid", fgColor=LIGHT_GRAY)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def style_cell(ws, row, col, value=None, font=None, fill=None, alignment=None,
               border=thin_border, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def style_row(ws, row, col_start, col_end, font=None, fill=None, alignment=None,
              border=thin_border, number_format=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format


def merge_and_style(ws, range_str, value, font, fill, alignment=align_center):
    ws.merge_cells(range_str)
    top_left = range_str.split(":")[0]
    cell = ws[top_left]
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment


# ============================================================
# MONTH HELPERS
# ============================================================
MONTHS = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
# Columns: A=1(Category), B=2(Jan)..M=13(Dec), N=14(YTD), O=15(% of Rev)
COL_CAT = 1
COL_JAN = 2  # B
COL_DEC = 13  # M
COL_YTD = 14  # N
COL_PCT = 15  # O
REVENUE_ROW = 6  # Total Revenue row


# ============================================================
# TAB 1: Monthly P&L
# ============================================================
ws1 = wb.active
ws1.title = "Monthly P&L"
ws1.sheet_properties.tabColor = DARK_GREEN

# Column widths
ws1.column_dimensions["A"].width = 28
for c in range(COL_JAN, COL_DEC + 1):
    ws1.column_dimensions[get_column_letter(c)].width = 12
ws1.column_dimensions[get_column_letter(COL_YTD)].width = 14
ws1.column_dimensions[get_column_letter(COL_PCT)].width = 12

# --- Row 1: Title (merged A1:O1) ---
merge_and_style(ws1, "A1:O1", "TOTALGUARD YARD CARE — 2026 PROFIT & LOSS",
                font_title, fill_title)

# --- Row 2: Headers ---
style_cell(ws1, 2, COL_CAT, "Category", font_header, fill_header, align_center)
for i, month in enumerate(MONTHS):
    style_cell(ws1, 2, COL_JAN + i, month, font_header, fill_header, align_center)
style_cell(ws1, 2, COL_YTD, "YTD Total", font_header, fill_header, align_center)
style_cell(ws1, 2, COL_PCT, "% of Revenue", font_header, fill_header, align_center)

# Freeze below headers
ws1.freeze_panes = "A3"


def month_cols():
    """Return column letters B through M."""
    return [get_column_letter(c) for c in range(COL_JAN, COL_DEC + 1)]


def ytd_formula(row):
    """SUM(B{row}:M{row})"""
    return f"=SUM(B{row}:M{row})"


def pct_formula(row):
    """=N{row}/N$6"""
    return f"=IF(N${REVENUE_ROW}=0,0,N{row}/N${REVENUE_ROW})"


def sum_formula_monthly(rows, col_letter):
    """=SUM of specific rows in a column."""
    refs = "+".join([f"{col_letter}{r}" for r in rows])
    return f"={refs}"


def diff_formula_monthly(row_a, row_b, col_letter):
    """=rowA - rowB in a column."""
    return f"={col_letter}{row_a}-{col_letter}{row_b}"


def section_header(ws, row, text, fill):
    """Merge A:O for a section header row."""
    merge_and_style(ws, f"A{row}:O{row}", text, font_section, fill, align_left)


def write_data_row(ws, row, label, row_type="manual", source_rows=None,
                   op="sum", alt_row_idx=0):
    """
    row_type: manual | sum | diff | pct_margin
    source_rows: list of row numbers for sum, or [a, b] for diff (a-b) or pct (a/b)
    """
    is_total = row_type in ("sum", "diff")
    is_pct = row_type == "pct_margin"
    is_bold = is_total
    is_big = label in ("NET INCOME",)

    f_font = font_total_big if is_big else (font_bold if is_bold else
              (font_italic if is_pct else font_normal))
    f_fill = fill_med_green if label in ("TOTAL REVENUE", "GROSS PROFIT", "NET INCOME") else (
             fill_white if alt_row_idx % 2 == 0 else fill_gray)
    num_fmt = "0.0%" if is_pct else '$#,##0.00'

    # Category label
    style_cell(ws, row, COL_CAT, label, f_font, f_fill, align_left)

    # Monthly columns B-M
    for c in range(COL_JAN, COL_DEC + 1):
        cl = get_column_letter(c)
        if row_type == "manual":
            style_cell(ws, row, c, None, font_normal, f_fill, align_right,
                       number_format=num_fmt)
        elif row_type == "sum":
            refs = "+".join([f"{cl}{r}" for r in source_rows])
            style_cell(ws, row, c, f"={refs}", f_font, f_fill, align_right,
                       number_format=num_fmt)
        elif row_type == "diff":
            style_cell(ws, row, c, f"={cl}{source_rows[0]}-{cl}{source_rows[1]}",
                       f_font, f_fill, align_right, number_format=num_fmt)
        elif row_type == "pct_margin":
            style_cell(ws, row, c,
                       f"=IF({cl}{source_rows[1]}=0,0,{cl}{source_rows[0]}/{cl}{source_rows[1]})",
                       f_font, f_fill, align_right, number_format=num_fmt)

    # YTD column N
    if is_pct:
        style_cell(ws, row, COL_YTD,
                   f"=IF(N{source_rows[1]}=0,0,N{source_rows[0]}/N{source_rows[1]})",
                   f_font, f_fill, align_right, number_format=num_fmt)
    else:
        style_cell(ws, row, COL_YTD, ytd_formula(row), f_font, f_fill, align_right,
                   number_format=num_fmt)

    # % of Revenue column O
    if is_pct:
        style_cell(ws, row, COL_PCT, None, f_font, f_fill, align_right)
    else:
        style_cell(ws, row, COL_PCT, pct_formula(row), f_font, f_fill, align_right,
                   number_format="0.0%")


# --- Row 3: REVENUE section header ---
section_header(ws1, 3, "REVENUE", fill_med_green)

# --- Rows 4-6: Revenue rows ---
write_data_row(ws1, 4, "Service Revenue", "manual", alt_row_idx=0)
write_data_row(ws1, 5, "Other Income", "manual", alt_row_idx=1)
write_data_row(ws1, 6, "TOTAL REVENUE", "sum", source_rows=[4, 5])

# --- Row 7: blank ---
# --- Row 8: COGS section header ---
section_header(ws1, 8, "COST OF GOODS SOLD", fill_gray)

# --- Rows 9-13: COGS items ---
write_data_row(ws1, 9, "Fuel & Gas", "manual", alt_row_idx=0)
write_data_row(ws1, 10, "Supplies & Materials", "manual", alt_row_idx=1)
write_data_row(ws1, 11, "Subcontractor Labor", "manual", alt_row_idx=0)
write_data_row(ws1, 12, "Equipment Rental", "manual", alt_row_idx=1)
write_data_row(ws1, 13, "TOTAL COGS", "sum", source_rows=[9, 10, 11, 12])

# --- Row 14: blank ---
# --- Row 15: Gross Profit ---
write_data_row(ws1, 15, "GROSS PROFIT", "diff", source_rows=[6, 13])

# --- Row 16: Gross Margin % ---
write_data_row(ws1, 16, "Gross Margin %", "pct_margin", source_rows=[15, 6])

# --- Row 17: blank ---
# --- Row 18: OPERATING EXPENSES section header ---
section_header(ws1, 18, "OPERATING EXPENSES", fill_gray)

# --- Rows 19-28: OpEx items ---
opex_items = [
    "Software/SaaS", "Phone/Communications", "Insurance",
    "Marketing/Advertising", "Vehicle Expenses", "Equipment Maintenance",
    "Office/Admin", "Licensing/Permits", "Training/Education", "Other Expenses"
]
for i, item in enumerate(opex_items):
    write_data_row(ws1, 19 + i, item, "manual", alt_row_idx=i)

# --- Row 29: TOTAL OPEX ---
write_data_row(ws1, 29, "TOTAL OPEX", "sum",
               source_rows=list(range(19, 29)))

# --- Row 30: blank ---
# --- Row 31: NET INCOME ---
write_data_row(ws1, 31, "NET INCOME", "diff", source_rows=[15, 29])

# --- Row 32: Net Margin % ---
write_data_row(ws1, 32, "Net Margin %", "pct_margin", source_rows=[31, 6])

# Print setup: landscape, fit to 1 page wide, repeat header rows 1-2
ws1.page_setup.orientation = "landscape"
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 0
ws1.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
    fitToPage=True
)
ws1.print_title_rows = "1:2"


# ============================================================
# TAB 2: Quarterly Summary
# ============================================================
ws2 = wb.create_sheet("Quarterly Summary")
ws2.sheet_properties.tabColor = DARK_GREEN

# Column widths: A=24, B-F=16
ws2.column_dimensions["A"].width = 24
for c in range(2, 7):
    ws2.column_dimensions[get_column_letter(c)].width = 16

# --- Row 1: Title (merged A1:F1) ---
merge_and_style(ws2, "A1:F1",
                "TOTALGUARD YARD CARE — 2026 QUARTERLY SUMMARY",
                font_title, fill_title)

# --- Row 2: Headers ---
q_headers = ["Metric", "Q1", "Q2", "Q3", "Q4", "Full Year"]
for i, h in enumerate(q_headers):
    style_cell(ws2, 2, i + 1, h, font_header, fill_header, align_center)

ws2.freeze_panes = "A3"

# Quarter month mappings (column letters on Monthly P&L tab)
# Q1 = B+C+D (Jan-Mar), Q2 = E+F+G (Apr-Jun), Q3 = H+I+J (Jul-Sep), Q4 = K+L+M (Oct-Dec)
Q_COLS = [
    ["B", "C", "D"],   # Q1
    ["E", "F", "G"],   # Q2
    ["H", "I", "J"],   # Q3
    ["K", "L", "M"],   # Q4
]

# Source rows on Monthly P&L
SRC = {
    "Revenue": 6,
    "COGS": 13,
    "Operating Expenses": 29,
}


def q_sum_formula(src_row, q_idx):
    """Sum 3 months from Monthly P&L for a quarter."""
    cols = Q_COLS[q_idx]
    refs = "+".join([f"'Monthly P&L'!{c}{src_row}" for c in cols])
    return f"={refs}"


def full_year_formula(row):
    """Sum B through E on this sheet (Q1-Q4)."""
    return f"=SUM(B{row}:E{row})"


# Quarterly rows
q_rows = [
    ("Revenue", 6, "currency", None),
    ("COGS", 13, "currency", None),
    ("Gross Profit", None, "currency", "diff"),      # Revenue - COGS
    ("Gross Margin %", None, "pct", "margin"),        # GP / Revenue
    ("Operating Expenses", 29, "currency", None),
    ("Net Income", None, "currency", "diff2"),        # GP - OpEx
    ("Net Margin %", None, "pct", "margin2"),         # NI / Revenue
]

for idx, (label, src_row, fmt, calc_type) in enumerate(q_rows):
    row = 3 + idx
    is_bold = label in ("Revenue", "Gross Profit", "Net Income",
                         "COGS", "Operating Expenses")
    is_pct = fmt == "pct"
    f_font = font_bold if is_bold else (font_italic if is_pct else font_normal)
    f_fill = fill_med_green if label in ("Gross Profit", "Net Income") else (
             fill_white if idx % 2 == 0 else fill_gray)
    num_fmt = "0.0%" if is_pct else '$#,##0.00'

    # Label
    style_cell(ws2, row, 1, label, f_font, f_fill, align_left)

    if src_row is not None:
        # Direct pull from Monthly P&L
        for qi in range(4):
            style_cell(ws2, row, qi + 2, q_sum_formula(src_row, qi),
                       f_font, f_fill, align_right, number_format=num_fmt)
        style_cell(ws2, row, 6, full_year_formula(row),
                   f_font, f_fill, align_right, number_format=num_fmt)
    elif calc_type == "diff":
        # Gross Profit = Revenue row - COGS row
        rev_r = 3  # Revenue is row 3
        cogs_r = 4  # COGS is row 4
        for c in range(2, 7):
            cl = get_column_letter(c)
            style_cell(ws2, row, c, f"={cl}{rev_r}-{cl}{cogs_r}",
                       f_font, f_fill, align_right, number_format=num_fmt)
    elif calc_type == "margin":
        # Gross Margin % = GP / Revenue
        gp_r = 5  # Gross Profit row
        rev_r = 3
        for c in range(2, 7):
            cl = get_column_letter(c)
            style_cell(ws2, row, c,
                       f"=IF({cl}{rev_r}=0,0,{cl}{gp_r}/{cl}{rev_r})",
                       f_font, f_fill, align_right, number_format=num_fmt)
    elif calc_type == "diff2":
        # Net Income = GP - OpEx
        gp_r = 5
        opex_r = 7  # Operating Expenses row
        for c in range(2, 7):
            cl = get_column_letter(c)
            style_cell(ws2, row, c, f"={cl}{gp_r}-{cl}{opex_r}",
                       f_font, f_fill, align_right, number_format=num_fmt)
    elif calc_type == "margin2":
        # Net Margin % = NI / Revenue
        ni_r = 8  # Net Income row
        rev_r = 3
        for c in range(2, 7):
            cl = get_column_letter(c)
            style_cell(ws2, row, c,
                       f"=IF({cl}{rev_r}=0,0,{cl}{ni_r}/{cl}{rev_r})",
                       f_font, f_fill, align_right, number_format=num_fmt)

# Print setup
ws2.page_setup.orientation = "landscape"
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
    fitToPage=True
)
ws2.print_title_rows = "1:2"


# ============================================================
# SAVE
# ============================================================
output_path = r"c:\Users\vance\OneDrive\Desktop\claude-workspace\TotalGuard_2026_PnL_Dashboard.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
