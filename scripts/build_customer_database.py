import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()

# ============================================================
# COLOR PALETTE (matches TotalGuard branding)
# ============================================================
DARK_GREEN = "1B4332"
LIGHT_GREEN_FILL = "D8F3DC"
MED_GREEN = "95D5B2"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
BORDER_COLOR = "CCCCCC"
BLACK = "000000"

# Conditional formatting colors
CF_ACTIVE = "D4EDDA"
CF_LEAD = "FFF3CD"
CF_INACTIVE = "E0E0E0"
CF_LOST = "F8D7DA"

# Reusable styles
font_title = Font(name="Arial", size=16, bold=True, color=DARK_GREEN)
font_header = Font(name="Arial", size=11, bold=True, color=WHITE)
font_normal = Font(name="Arial", size=10, color=BLACK)
font_bold = Font(name="Arial", size=10, bold=True, color=BLACK)
font_total = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
font_section = Font(name="Arial", size=12, bold=True, color=WHITE)

fill_header = PatternFill("solid", fgColor=DARK_GREEN)
fill_title = PatternFill("solid", fgColor=LIGHT_GREEN_FILL)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_gray = PatternFill("solid", fgColor=LIGHT_GRAY)
fill_total = PatternFill("solid", fgColor=MED_GREEN)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def style_row(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None, number_format=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format


# ============================================================
# TAB 1: CUSTOMERS
# ============================================================
ws1 = wb.active
ws1.title = "Customers"
ws1.sheet_properties.tabColor = DARK_GREEN

# Column widths: A-N
col_widths = {1: 8, 2: 14, 3: 14, 4: 16, 5: 26, 6: 28, 7: 14, 8: 28,
              9: 14, 10: 14, 11: 14, 12: 16, 13: 12, 14: 24}
for c, w in col_widths.items():
    ws1.column_dimensions[get_column_letter(c)].width = w

# Row 1: Title
ws1.merge_cells("A1:N1")
cell_title = ws1["A1"]
cell_title.value = "TOTALGUARD YARD CARE \u2014 2026 CUSTOMER DATABASE"
cell_title.font = font_title
cell_title.fill = fill_title
cell_title.alignment = align_center
# Apply title fill across merged range
for c in range(1, 15):
    ws1.cell(row=1, column=c).fill = fill_title

# Row 2: Headers
headers = ["Cust #", "First Name", "Last Name", "Phone", "Email", "Address",
           "City", "Services Used", "Customer Since", "Source",
           "Total Revenue", "Last Service Date", "Status", "Notes"]
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Data rows 3-502 (500 capacity)
for row in range(3, 503):
    alt_fill = fill_white if (row - 3) % 2 == 0 else fill_gray

    # A: Cust # (auto-increment 1-500)
    cell_num = ws1.cell(row=row, column=1, value=row - 2)
    cell_num.font = font_normal
    cell_num.fill = alt_fill
    cell_num.alignment = align_center
    cell_num.border = thin_border

    # B-N: Empty data cells with formatting
    for c in range(2, 15):
        cell = ws1.cell(row=row, column=c)
        cell.font = font_normal
        cell.fill = alt_fill
        cell.border = thin_border

        if c in (1,):  # Cust # already set
            cell.alignment = align_center
        elif c == 9:  # Customer Since
            cell.number_format = "MM/DD/YYYY"
            cell.alignment = align_center
        elif c == 11:  # Total Revenue
            cell.number_format = '$#,##0.00'
            cell.alignment = align_center
        elif c == 12:  # Last Service Date
            cell.number_format = "MM/DD/YYYY"
            cell.alignment = align_center
        elif c == 13:  # Status
            cell.alignment = align_center
        elif c in (7, 10):  # City, Source
            cell.alignment = align_center
        else:
            cell.alignment = align_left

# Row 503: TOTAL row
total_row = 503

# Style all cells first (before merging)
for c in range(1, 15):
    cell = ws1.cell(row=total_row, column=c)
    cell.fill = fill_total
    cell.border = thin_border
    cell.font = font_total

# Set values before merging
ws1.cell(row=total_row, column=1).value = "TOTAL"
ws1.cell(row=total_row, column=1).alignment = align_center

# Customer count in B503
ws1.cell(row=total_row, column=2).value = "=COUNTA(B3:B502)"
ws1.cell(row=total_row, column=2).number_format = "0"
ws1.cell(row=total_row, column=2).alignment = align_center

# Label in C
ws1.cell(row=total_row, column=3).value = "customers"
ws1.cell(row=total_row, column=3).alignment = align_left

# Total Revenue in K503
ws1.cell(row=total_row, column=11).value = "=SUM(K3:K502)"
ws1.cell(row=total_row, column=11).number_format = '$#,##0.00'
ws1.cell(row=total_row, column=11).alignment = align_center

# Data Validations
cities = "Madison,Middleton,Waunakee,Monona,Sun Prairie,Fitchburg,Verona,McFarland,Cottage Grove,DeForest,Oregon,Stoughton"
dv_city = DataValidation(type="list", formula1=f'"{cities}"', allow_blank=True)
dv_city.error = "Please select a valid city"
dv_city.errorTitle = "Invalid City"
ws1.add_data_validation(dv_city)
dv_city.add(f"G3:G502")

sources = "Google,Referral,Yard Sign,Door Knock,Facebook,Nextdoor,Website,Jobber,Repeat,Other"
dv_source = DataValidation(type="list", formula1=f'"{sources}"', allow_blank=True)
dv_source.error = "Please select a valid source"
dv_source.errorTitle = "Invalid Source"
ws1.add_data_validation(dv_source)
dv_source.add(f"J3:J502")

statuses = "Active,Inactive,Lead,Lost"
dv_status = DataValidation(type="list", formula1=f'"{statuses}"', allow_blank=True)
dv_status.error = "Please select a valid status"
dv_status.errorTitle = "Invalid Status"
ws1.add_data_validation(dv_status)
dv_status.add(f"M3:M502")

# Conditional formatting on column M (Status)
ws1.conditional_formatting.add(
    "M3:M502",
    CellIsRule(operator="equal", formula=['"Active"'],
              fill=PatternFill("solid", fgColor=CF_ACTIVE))
)
ws1.conditional_formatting.add(
    "M3:M502",
    CellIsRule(operator="equal", formula=['"Lead"'],
              fill=PatternFill("solid", fgColor=CF_LEAD))
)
ws1.conditional_formatting.add(
    "M3:M502",
    CellIsRule(operator="equal", formula=['"Inactive"'],
              fill=PatternFill("solid", fgColor=CF_INACTIVE))
)
ws1.conditional_formatting.add(
    "M3:M502",
    CellIsRule(operator="equal", formula=['"Lost"'],
              fill=PatternFill("solid", fgColor=CF_LOST))
)

# Freeze panes below headers (row 2)
ws1.freeze_panes = "A3"

# Print setup
ws1.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws1.page_setup.orientation = "landscape"
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 0
ws1.print_title_rows = "1:2"


# ============================================================
# TAB 2: CUSTOMER SUMMARY
# ============================================================
ws2 = wb.create_sheet("Customer Summary")
ws2.sheet_properties.tabColor = DARK_GREEN

# Column widths
ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 14

# Row 1: Title
ws2.merge_cells("A1:B1")
cell_title2 = ws2["A1"]
cell_title2.value = "TOTALGUARD YARD CARE \u2014 2026 CUSTOMER SUMMARY"
cell_title2.font = font_title
cell_title2.fill = fill_title
cell_title2.alignment = align_center
ws2.cell(row=1, column=2).fill = fill_title

# Row 2: Headers
for i, h in enumerate(["Metric", "Value"], 1):
    cell = ws2.cell(row=2, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Helper to add a summary row
def add_summary_row(ws, row, label, formula, is_currency=False, fill=None):
    alt = fill if fill else (fill_white if (row - 3) % 2 == 0 else fill_gray)
    cell_a = ws.cell(row=row, column=1, value=label)
    cell_a.font = font_bold
    cell_a.fill = alt
    cell_a.alignment = align_left
    cell_a.border = thin_border

    cell_b = ws.cell(row=row, column=2, value=formula)
    cell_b.font = font_bold
    cell_b.fill = alt
    cell_b.alignment = align_center
    cell_b.border = thin_border
    if is_currency:
        cell_b.number_format = '$#,##0.00'


# Rows 3-7: Status counts
add_summary_row(ws2, 3, "Total Customers", "=COUNTA(Customers!B$3:B$502)")
add_summary_row(ws2, 4, "Active Customers", '=COUNTIF(Customers!M$3:M$502,"Active")')
add_summary_row(ws2, 5, "Leads", '=COUNTIF(Customers!M$3:M$502,"Lead")')
add_summary_row(ws2, 6, "Inactive", '=COUNTIF(Customers!M$3:M$502,"Inactive")')
add_summary_row(ws2, 7, "Lost", '=COUNTIF(Customers!M$3:M$502,"Lost")')

# Row 8: blank
for c in range(1, 3):
    ws2.cell(row=8, column=c).border = thin_border

# Row 9-10: Revenue
add_summary_row(ws2, 9, "Total Revenue", "=SUM(Customers!K$3:K$502)", is_currency=True)
add_summary_row(ws2, 10, "Avg Revenue/Customer", "=IF(B3>0,B9/B3,0)", is_currency=True)

# Row 11: blank
for c in range(1, 3):
    ws2.cell(row=11, column=c).border = thin_border

# ---- Section: Customers by City ----
# Row 12: section header
for i, h in enumerate(["City", "Count"], 1):
    cell = ws2.cell(row=12, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

city_list = ["Madison", "Middleton", "Waunakee", "Monona", "Sun Prairie",
             "Fitchburg", "Verona", "McFarland", "Cottage Grove", "DeForest",
             "Oregon", "Stoughton"]

for idx, city in enumerate(city_list):
    row = 13 + idx
    alt = fill_white if idx % 2 == 0 else fill_gray
    cell_a = ws2.cell(row=row, column=1, value=city)
    cell_a.font = font_normal
    cell_a.fill = alt
    cell_a.alignment = align_left
    cell_a.border = thin_border

    cell_b = ws2.cell(row=row, column=2, value=f'=COUNTIF(Customers!G$3:G$502,"{city}")')
    cell_b.font = font_normal
    cell_b.fill = alt
    cell_b.alignment = align_center
    cell_b.border = thin_border

# Row 25: City total
city_total_row = 25
cell_a = ws2.cell(row=city_total_row, column=1, value="Total")
cell_a.font = font_total
cell_a.fill = fill_total
cell_a.alignment = align_left
cell_a.border = thin_border

cell_b = ws2.cell(row=city_total_row, column=2, value="=SUM(B13:B24)")
cell_b.font = font_total
cell_b.fill = fill_total
cell_b.alignment = align_center
cell_b.border = thin_border

# Row 26: blank
for c in range(1, 3):
    ws2.cell(row=26, column=c).border = thin_border

# ---- Section: Customers by Source ----
# Row 27: section header
for i, h in enumerate(["Source", "Count"], 1):
    cell = ws2.cell(row=27, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

source_list = ["Google", "Referral", "Yard Sign", "Door Knock", "Facebook",
               "Nextdoor", "Website", "Jobber", "Repeat", "Other"]

for idx, source in enumerate(source_list):
    row = 28 + idx
    alt = fill_white if idx % 2 == 0 else fill_gray
    cell_a = ws2.cell(row=row, column=1, value=source)
    cell_a.font = font_normal
    cell_a.fill = alt
    cell_a.alignment = align_left
    cell_a.border = thin_border

    cell_b = ws2.cell(row=row, column=2, value=f'=COUNTIF(Customers!J$3:J$502,"{source}")')
    cell_b.font = font_normal
    cell_b.fill = alt
    cell_b.alignment = align_center
    cell_b.border = thin_border

# Row 38: Source total
source_total_row = 38
cell_a = ws2.cell(row=source_total_row, column=1, value="Total")
cell_a.font = font_total
cell_a.fill = fill_total
cell_a.alignment = align_left
cell_a.border = thin_border

cell_b = ws2.cell(row=source_total_row, column=2, value="=SUM(B28:B37)")
cell_b.font = font_total
cell_b.fill = fill_total
cell_b.alignment = align_center
cell_b.border = thin_border

# Freeze panes below headers
ws2.freeze_panes = "A3"

# Print setup
ws2.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws2.page_setup.orientation = "landscape"
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.print_title_rows = "1:2"


# ============================================================
# SAVE
# ============================================================
output_path = r"c:\Users\vance\OneDrive\Desktop\claude-workspace\TotalGuard_2026_Customer_Database.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
