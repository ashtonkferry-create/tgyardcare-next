import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

wb = Workbook()

# ============================================================
# COLOR PALETTE (matches TotalGuard branding)
# ============================================================
DARK_GREEN = "1B4332"
LIGHT_GREEN = "D8F3DC"
MED_GREEN = "95D5B2"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
BORDER_COLOR = "CCCCCC"
BLACK = "000000"

# Conditional formatting fills
CF_GREEN = "D4EDDA"
CF_YELLOW = "FFF3CD"
CF_ORANGE = "FFE0B2"
CF_RED = "F8D7DA"
CF_BLUE = "CCE5FF"
CF_GRAY_FILL = "E0E0E0"

# Reusable styles
font_title = Font(name="Arial", size=16, bold=True, color=DARK_GREEN)
font_header = Font(name="Arial", size=11, bold=True, color=WHITE)
font_normal = Font(name="Arial", size=10, color=BLACK)
font_bold = Font(name="Arial", size=10, bold=True, color=BLACK)
font_total = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)

fill_header = PatternFill("solid", fgColor=DARK_GREEN)
fill_title = PatternFill("solid", fgColor=LIGHT_GREEN)
fill_total = PatternFill("solid", fgColor=MED_GREEN)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_gray = PatternFill("solid", fgColor=LIGHT_GRAY)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def style_row(ws, row, col_range, font=None, fill=None, alignment=None, border=None, number_format=None):
    for c in col_range:
        cell = ws.cell(row=row, column=c)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format


# ============================================================
# TAB 1: OUTSTANDING INVOICES
# ============================================================
ws1 = wb.active
ws1.title = "Outstanding Invoices"
ws1.sheet_properties.tabColor = DARK_GREEN

# Column widths: A-L
col_widths = {1: 10, 2: 24, 3: 20, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 14, 10: 14, 11: 14, 12: 24}
for c, w in col_widths.items():
    ws1.column_dimensions[get_column_letter(c)].width = w

# Row 1: Title
ws1.merge_cells("A1:L1")
cell_title = ws1["A1"]
cell_title.value = "TOTALGUARD YARD CARE \u2014 2026 ACCOUNTS RECEIVABLE"
cell_title.font = font_title
cell_title.fill = fill_title
cell_title.alignment = align_center
# Apply fill to all merged cells in row 1
for col in range(2, 13):
    ws1.cell(row=1, column=col).fill = fill_title

# Row 2: Headers
headers = [
    "Invoice #", "Customer", "Service", "Invoice Date", "Due Date",
    "Amount", "Amount Paid", "Balance Due", "Days Outstanding",
    "Status", "Last Follow-up", "Notes"
]
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Data rows: 3-202 (200 rows)
# Service dropdown
service_list = (
    "Lawn Mowing,Fertilization,Aeration,Herbicide,Weeding,Mulching,"
    "Garden Bed Care,Bush Trimming,Hardscaping,Spring Cleanup,"
    "Fall Cleanup,Leaf Removal,Gutter Cleaning,Gutter Guard Installation,Snow Removal"
)
dv_service = DataValidation(type="list", formula1=f'"{service_list}"', allow_blank=True)
dv_service.error = "Please select a valid service."
dv_service.errorTitle = "Invalid Service"
ws1.add_data_validation(dv_service)

# Status dropdown
status_list = "Current,30 Days,60 Days,90+ Days,Paid,Written Off"
dv_status = DataValidation(type="list", formula1=f'"{status_list}"', allow_blank=True)
dv_status.error = "Please select a valid status."
dv_status.errorTitle = "Invalid Status"
ws1.add_data_validation(dv_status)

for row in range(3, 203):
    fill = fill_white if (row - 3) % 2 == 0 else fill_gray

    # A: Invoice # (auto-increment)
    cell_inv = ws1.cell(row=row, column=1, value=row - 2)
    cell_inv.font = font_normal
    cell_inv.fill = fill
    cell_inv.alignment = align_center
    cell_inv.border = thin_border

    # B: Customer
    cell_cust = ws1.cell(row=row, column=2)
    cell_cust.font = font_normal
    cell_cust.fill = fill
    cell_cust.alignment = align_left
    cell_cust.border = thin_border

    # C: Service (dropdown)
    cell_svc = ws1.cell(row=row, column=3)
    cell_svc.font = font_normal
    cell_svc.fill = fill
    cell_svc.alignment = align_center
    cell_svc.border = thin_border
    dv_service.add(cell_svc)

    # D: Invoice Date
    cell_inv_date = ws1.cell(row=row, column=4)
    cell_inv_date.font = font_normal
    cell_inv_date.fill = fill
    cell_inv_date.alignment = align_center
    cell_inv_date.border = thin_border
    cell_inv_date.number_format = "MM/DD/YYYY"

    # E: Due Date
    cell_due = ws1.cell(row=row, column=5)
    cell_due.font = font_normal
    cell_due.fill = fill
    cell_due.alignment = align_center
    cell_due.border = thin_border
    cell_due.number_format = "MM/DD/YYYY"

    # F: Amount
    cell_amt = ws1.cell(row=row, column=6)
    cell_amt.font = font_normal
    cell_amt.fill = fill
    cell_amt.alignment = align_center
    cell_amt.border = thin_border
    cell_amt.number_format = "$#,##0.00"

    # G: Amount Paid
    cell_paid = ws1.cell(row=row, column=7)
    cell_paid.font = font_normal
    cell_paid.fill = fill
    cell_paid.alignment = align_center
    cell_paid.border = thin_border
    cell_paid.number_format = "$#,##0.00"

    # H: Balance Due = F - G
    cell_bal = ws1.cell(row=row, column=8, value=f"=F{row}-G{row}")
    cell_bal.font = font_normal
    cell_bal.fill = fill
    cell_bal.alignment = align_center
    cell_bal.border = thin_border
    cell_bal.number_format = "$#,##0.00"

    # I: Days Outstanding = IF(AND(D{row}<>"",H{row}>0),TODAY()-D{row},"")
    cell_days = ws1.cell(row=row, column=9, value=f'=IF(AND(D{row}<>"",H{row}>0),TODAY()-D{row},"")')
    cell_days.font = font_normal
    cell_days.fill = fill
    cell_days.alignment = align_center
    cell_days.border = thin_border
    cell_days.number_format = "0"

    # J: Status (dropdown)
    cell_status = ws1.cell(row=row, column=10)
    cell_status.font = font_normal
    cell_status.fill = fill
    cell_status.alignment = align_center
    cell_status.border = thin_border
    dv_status.add(cell_status)

    # K: Last Follow-up
    cell_followup = ws1.cell(row=row, column=11)
    cell_followup.font = font_normal
    cell_followup.fill = fill
    cell_followup.alignment = align_center
    cell_followup.border = thin_border
    cell_followup.number_format = "MM/DD/YYYY"

    # L: Notes
    cell_notes = ws1.cell(row=row, column=12)
    cell_notes.font = font_normal
    cell_notes.fill = fill
    cell_notes.alignment = align_left
    cell_notes.border = thin_border

# Row 203: TOTAL row
total_row = 203
ws1.merge_cells(f"A{total_row}:E{total_row}")
cell_total_label = ws1[f"A{total_row}"]
cell_total_label.value = "TOTAL"
cell_total_label.font = Font(name="Arial", size=12, bold=True, color=DARK_GREEN)
cell_total_label.fill = fill_total
cell_total_label.alignment = align_center
cell_total_label.border = thin_border
for col in range(2, 6):
    ws1.cell(row=total_row, column=col).fill = fill_total
    ws1.cell(row=total_row, column=col).border = thin_border

# SUM for Amount (F), Amount Paid (G), Balance Due (H)
for col_letter in ["F", "G", "H"]:
    col_idx = ord(col_letter) - ord("A") + 1
    cell = ws1.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}3:{col_letter}202)")
    cell.font = Font(name="Arial", size=12, bold=True, color=DARK_GREEN)
    cell.fill = fill_total
    cell.alignment = align_center
    cell.border = thin_border
    cell.number_format = "$#,##0.00"

# Fill remaining total row cells
for col in range(9, 13):
    ws1.cell(row=total_row, column=col).fill = fill_total
    ws1.cell(row=total_row, column=col).border = thin_border

# ---- Conditional Formatting on Status (Column J) ----
status_range = "J3:J202"

ws1.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"Current"'], fill=PatternFill("solid", fgColor=CF_GREEN))
)
ws1.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"30 Days"'], fill=PatternFill("solid", fgColor=CF_YELLOW))
)
ws1.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"60 Days"'], fill=PatternFill("solid", fgColor=CF_ORANGE))
)
ws1.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"90+ Days"'], fill=PatternFill("solid", fgColor=CF_RED))
)
ws1.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"Paid"'], fill=PatternFill("solid", fgColor=CF_BLUE))
)
ws1.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"Written Off"'], fill=PatternFill("solid", fgColor=CF_GRAY_FILL))
)

# ---- Conditional Formatting on Days Outstanding (Column I) ----
days_range = "I3:I202"

# > 60 = red text (check this first so >60 takes priority over >30)
ws1.conditional_formatting.add(
    days_range,
    CellIsRule(operator="greaterThan", formula=["60"], font=Font(name="Arial", size=10, color="CC0000"))
)
# > 30 = orange text
ws1.conditional_formatting.add(
    days_range,
    CellIsRule(operator="greaterThan", formula=["30"], font=Font(name="Arial", size=10, color="E65100"))
)

# Freeze panes below header row 2
ws1.freeze_panes = "A3"

# Print setup
ws1.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws1.page_setup.orientation = "landscape"
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 0
ws1.print_title_rows = "1:2"


# ============================================================
# TAB 2: AGING SUMMARY
# ============================================================
ws2 = wb.create_sheet("Aging Summary")
ws2.sheet_properties.tabColor = MED_GREEN

col_widths_2 = {1: 28, 2: 12, 3: 18, 4: 14}
for c, w in col_widths_2.items():
    ws2.column_dimensions[get_column_letter(c)].width = w

# Row 1: Title
ws2.merge_cells("A1:D1")
cell_title2 = ws2["A1"]
cell_title2.value = "TOTALGUARD YARD CARE \u2014 2026 AR AGING SUMMARY"
cell_title2.font = font_title
cell_title2.fill = fill_title
cell_title2.alignment = align_center
for col in range(2, 5):
    ws2.cell(row=1, column=col).fill = fill_title

# Row 2: Headers
aging_headers = ["Category", "Count", "Amount", "% of Total"]
for i, h in enumerate(aging_headers, 1):
    cell = ws2.cell(row=2, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Reference sheet name for formulas
inv_sheet = "'Outstanding Invoices'"

# Data rows start at 3
# Row 3: Current (0-30 days)
r = 3
ws2.cell(row=r, column=1, value="Current (0-30 days)").font = font_bold
ws2.cell(row=r, column=1).fill = fill_white
ws2.cell(row=r, column=1).alignment = align_left
ws2.cell(row=r, column=1).border = thin_border

ws2.cell(row=r, column=2, value=f'=COUNTIF({inv_sheet}!J3:J202,"Current")')
ws2.cell(row=r, column=2).font = font_normal
ws2.cell(row=r, column=2).fill = fill_white
ws2.cell(row=r, column=2).alignment = align_center
ws2.cell(row=r, column=2).border = thin_border
ws2.cell(row=r, column=2).number_format = "0"

ws2.cell(row=r, column=3, value=f'=SUMIF({inv_sheet}!J3:J202,"Current",{inv_sheet}!H3:H202)')
ws2.cell(row=r, column=3).font = font_normal
ws2.cell(row=r, column=3).fill = fill_white
ws2.cell(row=r, column=3).alignment = align_center
ws2.cell(row=r, column=3).border = thin_border
ws2.cell(row=r, column=3).number_format = "$#,##0.00"

ws2.cell(row=r, column=4, value=f"=IF(C7=0,0,C{r}/C7)")
ws2.cell(row=r, column=4).font = font_normal
ws2.cell(row=r, column=4).fill = fill_white
ws2.cell(row=r, column=4).alignment = align_center
ws2.cell(row=r, column=4).border = thin_border
ws2.cell(row=r, column=4).number_format = "0.0%"

# Row 4: 31-60 Days
r = 4
ws2.cell(row=r, column=1, value="31-60 Days").font = font_bold
ws2.cell(row=r, column=1).fill = fill_gray
ws2.cell(row=r, column=1).alignment = align_left
ws2.cell(row=r, column=1).border = thin_border

ws2.cell(row=r, column=2, value=f'=COUNTIF({inv_sheet}!J3:J202,"30 Days")')
ws2.cell(row=r, column=2).font = font_normal
ws2.cell(row=r, column=2).fill = fill_gray
ws2.cell(row=r, column=2).alignment = align_center
ws2.cell(row=r, column=2).border = thin_border
ws2.cell(row=r, column=2).number_format = "0"

ws2.cell(row=r, column=3, value=f'=SUMIF({inv_sheet}!J3:J202,"30 Days",{inv_sheet}!H3:H202)')
ws2.cell(row=r, column=3).font = font_normal
ws2.cell(row=r, column=3).fill = fill_gray
ws2.cell(row=r, column=3).alignment = align_center
ws2.cell(row=r, column=3).border = thin_border
ws2.cell(row=r, column=3).number_format = "$#,##0.00"

ws2.cell(row=r, column=4, value=f"=IF(C7=0,0,C{r}/C7)")
ws2.cell(row=r, column=4).font = font_normal
ws2.cell(row=r, column=4).fill = fill_gray
ws2.cell(row=r, column=4).alignment = align_center
ws2.cell(row=r, column=4).border = thin_border
ws2.cell(row=r, column=4).number_format = "0.0%"

# Row 5: 61-90 Days
r = 5
ws2.cell(row=r, column=1, value="61-90 Days").font = font_bold
ws2.cell(row=r, column=1).fill = fill_white
ws2.cell(row=r, column=1).alignment = align_left
ws2.cell(row=r, column=1).border = thin_border

ws2.cell(row=r, column=2, value=f'=COUNTIF({inv_sheet}!J3:J202,"60 Days")')
ws2.cell(row=r, column=2).font = font_normal
ws2.cell(row=r, column=2).fill = fill_white
ws2.cell(row=r, column=2).alignment = align_center
ws2.cell(row=r, column=2).border = thin_border
ws2.cell(row=r, column=2).number_format = "0"

ws2.cell(row=r, column=3, value=f'=SUMIF({inv_sheet}!J3:J202,"60 Days",{inv_sheet}!H3:H202)')
ws2.cell(row=r, column=3).font = font_normal
ws2.cell(row=r, column=3).fill = fill_white
ws2.cell(row=r, column=3).alignment = align_center
ws2.cell(row=r, column=3).border = thin_border
ws2.cell(row=r, column=3).number_format = "$#,##0.00"

ws2.cell(row=r, column=4, value=f"=IF(C7=0,0,C{r}/C7)")
ws2.cell(row=r, column=4).font = font_normal
ws2.cell(row=r, column=4).fill = fill_white
ws2.cell(row=r, column=4).alignment = align_center
ws2.cell(row=r, column=4).border = thin_border
ws2.cell(row=r, column=4).number_format = "0.0%"

# Row 6: 90+ Days
r = 6
ws2.cell(row=r, column=1, value="90+ Days").font = font_bold
ws2.cell(row=r, column=1).fill = fill_gray
ws2.cell(row=r, column=1).alignment = align_left
ws2.cell(row=r, column=1).border = thin_border

ws2.cell(row=r, column=2, value=f'=COUNTIF({inv_sheet}!J3:J202,"90+ Days")')
ws2.cell(row=r, column=2).font = font_normal
ws2.cell(row=r, column=2).fill = fill_gray
ws2.cell(row=r, column=2).alignment = align_center
ws2.cell(row=r, column=2).border = thin_border
ws2.cell(row=r, column=2).number_format = "0"

ws2.cell(row=r, column=3, value=f'=SUMIF({inv_sheet}!J3:J202,"90+ Days",{inv_sheet}!H3:H202)')
ws2.cell(row=r, column=3).font = font_normal
ws2.cell(row=r, column=3).fill = fill_gray
ws2.cell(row=r, column=3).alignment = align_center
ws2.cell(row=r, column=3).border = thin_border
ws2.cell(row=r, column=3).number_format = "$#,##0.00"

ws2.cell(row=r, column=4, value=f"=IF(C7=0,0,C{r}/C7)")
ws2.cell(row=r, column=4).font = font_normal
ws2.cell(row=r, column=4).fill = fill_gray
ws2.cell(row=r, column=4).alignment = align_center
ws2.cell(row=r, column=4).border = thin_border
ws2.cell(row=r, column=4).number_format = "0.0%"

# Row 7: Total Outstanding
r = 7
ws2.cell(row=r, column=1, value="Total Outstanding").font = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
ws2.cell(row=r, column=1).fill = fill_total
ws2.cell(row=r, column=1).alignment = align_left
ws2.cell(row=r, column=1).border = thin_border

ws2.cell(row=r, column=2, value="=SUM(B3:B6)")
ws2.cell(row=r, column=2).font = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
ws2.cell(row=r, column=2).fill = fill_total
ws2.cell(row=r, column=2).alignment = align_center
ws2.cell(row=r, column=2).border = thin_border
ws2.cell(row=r, column=2).number_format = "0"

ws2.cell(row=r, column=3, value="=SUM(C3:C6)")
ws2.cell(row=r, column=3).font = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
ws2.cell(row=r, column=3).fill = fill_total
ws2.cell(row=r, column=3).alignment = align_center
ws2.cell(row=r, column=3).border = thin_border
ws2.cell(row=r, column=3).number_format = "$#,##0.00"

ws2.cell(row=r, column=4, value="=IF(C7=0,0,SUM(D3:D6))")
ws2.cell(row=r, column=4).font = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
ws2.cell(row=r, column=4).fill = fill_total
ws2.cell(row=r, column=4).alignment = align_center
ws2.cell(row=r, column=4).border = thin_border
ws2.cell(row=r, column=4).number_format = "0.0%"

# Row 8: blank
r = 8
for col in range(1, 5):
    ws2.cell(row=r, column=col).border = thin_border

# Row 9: Paid / Collected
r = 9
ws2.cell(row=r, column=1, value="Paid / Collected").font = font_bold
ws2.cell(row=r, column=1).fill = fill_white
ws2.cell(row=r, column=1).alignment = align_left
ws2.cell(row=r, column=1).border = thin_border

ws2.cell(row=r, column=2, value=f'=COUNTIF({inv_sheet}!J3:J202,"Paid")')
ws2.cell(row=r, column=2).font = font_normal
ws2.cell(row=r, column=2).fill = fill_white
ws2.cell(row=r, column=2).alignment = align_center
ws2.cell(row=r, column=2).border = thin_border
ws2.cell(row=r, column=2).number_format = "0"

ws2.cell(row=r, column=3, value=f'=SUMIF({inv_sheet}!J3:J202,"Paid",{inv_sheet}!H3:H202)')
ws2.cell(row=r, column=3).font = font_normal
ws2.cell(row=r, column=3).fill = fill_white
ws2.cell(row=r, column=3).alignment = align_center
ws2.cell(row=r, column=3).border = thin_border
ws2.cell(row=r, column=3).number_format = "$#,##0.00"

ws2.cell(row=r, column=4).fill = fill_white
ws2.cell(row=r, column=4).border = thin_border

# Row 10: Written Off
r = 10
ws2.cell(row=r, column=1, value="Written Off").font = font_bold
ws2.cell(row=r, column=1).fill = fill_gray
ws2.cell(row=r, column=1).alignment = align_left
ws2.cell(row=r, column=1).border = thin_border

ws2.cell(row=r, column=2, value=f'=COUNTIF({inv_sheet}!J3:J202,"Written Off")')
ws2.cell(row=r, column=2).font = font_normal
ws2.cell(row=r, column=2).fill = fill_gray
ws2.cell(row=r, column=2).alignment = align_center
ws2.cell(row=r, column=2).border = thin_border
ws2.cell(row=r, column=2).number_format = "0"

ws2.cell(row=r, column=3, value=f'=SUMIF({inv_sheet}!J3:J202,"Written Off",{inv_sheet}!H3:H202)')
ws2.cell(row=r, column=3).font = font_normal
ws2.cell(row=r, column=3).fill = fill_gray
ws2.cell(row=r, column=3).alignment = align_center
ws2.cell(row=r, column=3).border = thin_border
ws2.cell(row=r, column=3).number_format = "$#,##0.00"

ws2.cell(row=r, column=4).fill = fill_gray
ws2.cell(row=r, column=4).border = thin_border

# Row 11: blank
r = 11
for col in range(1, 5):
    ws2.cell(row=r, column=col).border = thin_border

# Row 12: GRAND TOTAL (All Invoices)
r = 12
ws2.cell(row=r, column=1, value="GRAND TOTAL (All Invoices)").font = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
ws2.cell(row=r, column=1).fill = fill_total
ws2.cell(row=r, column=1).alignment = align_left
ws2.cell(row=r, column=1).border = thin_border

ws2.cell(row=r, column=2, value=f"=B7+B9+B10")
ws2.cell(row=r, column=2).font = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
ws2.cell(row=r, column=2).fill = fill_total
ws2.cell(row=r, column=2).alignment = align_center
ws2.cell(row=r, column=2).border = thin_border
ws2.cell(row=r, column=2).number_format = "0"

ws2.cell(row=r, column=3, value=f"=SUM({inv_sheet}!F3:F202)")
ws2.cell(row=r, column=3).font = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
ws2.cell(row=r, column=3).fill = fill_total
ws2.cell(row=r, column=3).alignment = align_center
ws2.cell(row=r, column=3).border = thin_border
ws2.cell(row=r, column=3).number_format = "$#,##0.00"

ws2.cell(row=r, column=4).fill = fill_total
ws2.cell(row=r, column=4).border = thin_border

# Freeze panes below headers
ws2.freeze_panes = "A3"

# Print setup
ws2.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws2.page_setup.orientation = "landscape"
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.print_title_rows = "1:2"

# ============================================================
# SAVE
# ============================================================
output_path = r"c:\Users\vance\OneDrive\Desktop\claude-workspace\TotalGuard_2026_Accounts_Receivable.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
