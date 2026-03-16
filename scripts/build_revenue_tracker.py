import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()

# ============================================================
# COLOR PALETTE (matches TotalGuard branding)
# ============================================================
DARK_GREEN = "1B4332"
LIGHT_GREEN_TITLE = "D8F3DC"
MED_GREEN = "95D5B2"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
BORDER_COLOR = "CCCCCC"
BLACK = "000000"

# Conditional formatting fills
PAID_GREEN = "D4EDDA"
PENDING_YELLOW = "FFF3CD"
OVERDUE_RED = "F8D7DA"
PARTIAL_BLUE = "CCE5FF"

# Reusable styles
font_title = Font(name="Arial", size=16, bold=True, color=DARK_GREEN)
font_header = Font(name="Arial", size=11, bold=True, color=WHITE)
font_normal = Font(name="Arial", size=10, color=BLACK)
font_bold = Font(name="Arial", size=10, bold=True, color=BLACK)
font_total = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)

fill_header = PatternFill("solid", fgColor=DARK_GREEN)
fill_title = PatternFill("solid", fgColor=LIGHT_GREEN_TITLE)
fill_total = PatternFill("solid", fgColor=MED_GREEN)
fill_white = PatternFill("solid", fgColor=WHITE)
fill_gray = PatternFill("solid", fgColor=LIGHT_GRAY)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


def style_row(ws, row, col_range, font=None, fill=None, alignment=None, border=None, number_format=None):
    for c in col_range:
        cell = ws.cell(row=row, column=c)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format


def merge_and_style(ws, range_str, value, font, fill, alignment=align_center):
    ws.merge_cells(range_str)
    top_left = range_str.split(":")[0]
    cell = ws[top_left]
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment


# Service list used across tabs
SERVICES = [
    "Lawn Mowing", "Fertilization", "Aeration", "Herbicide", "Weeding",
    "Mulching", "Garden Bed Care", "Bush Trimming", "Hardscaping",
    "Spring Cleanup", "Fall Cleanup", "Leaf Removal", "Gutter Cleaning",
    "Gutter Guard Installation", "Snow Removal",
]

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

# ============================================================
# TAB 1: REVENUE LOG
# ============================================================
ws1 = wb.active
ws1.title = "Revenue Log"
ws1.sheet_properties.tabColor = DARK_GREEN

# Column widths
col_widths = {1: 6, 2: 14, 3: 24, 4: 22, 5: 28, 6: 16, 7: 16, 8: 16, 9: 14, 10: 14, 11: 24}
for c, w in col_widths.items():
    ws1.column_dimensions[get_column_letter(c)].width = w

# Row 1: Title
merge_and_style(ws1, "A1:K1", "TOTALGUARD YARD CARE \u2014 2026 REVENUE TRACKER", font_title, fill_title)

# Row 2: Headers
headers = ["#", "Date", "Customer", "Service", "Address", "Amount Charged",
           "Amount Received", "Payment Method", "Payment Status", "Date Received", "Notes"]
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Data rows 3-502
for row in range(3, 503):
    fill = fill_white if (row - 3) % 2 == 0 else fill_gray

    # Column A: auto-increment number
    cell_num = ws1.cell(row=row, column=1, value=row - 2)
    cell_num.font = font_normal
    cell_num.fill = fill
    cell_num.alignment = align_center
    cell_num.border = thin_border

    # Column B: Date
    cell_date = ws1.cell(row=row, column=2)
    cell_date.font = font_normal
    cell_date.fill = fill
    cell_date.alignment = align_center
    cell_date.border = thin_border
    cell_date.number_format = "MM/DD/YYYY"

    # Column C: Customer
    ws1.cell(row=row, column=3).font = font_normal
    ws1.cell(row=row, column=3).fill = fill
    ws1.cell(row=row, column=3).alignment = align_left
    ws1.cell(row=row, column=3).border = thin_border

    # Column D: Service (dropdown applied later)
    ws1.cell(row=row, column=4).font = font_normal
    ws1.cell(row=row, column=4).fill = fill
    ws1.cell(row=row, column=4).alignment = align_left
    ws1.cell(row=row, column=4).border = thin_border

    # Column E: Address
    ws1.cell(row=row, column=5).font = font_normal
    ws1.cell(row=row, column=5).fill = fill
    ws1.cell(row=row, column=5).alignment = align_left
    ws1.cell(row=row, column=5).border = thin_border

    # Column F: Amount Charged
    ws1.cell(row=row, column=6).font = font_normal
    ws1.cell(row=row, column=6).fill = fill
    ws1.cell(row=row, column=6).alignment = align_right
    ws1.cell(row=row, column=6).border = thin_border
    ws1.cell(row=row, column=6).number_format = '$#,##0.00'

    # Column G: Amount Received
    ws1.cell(row=row, column=7).font = font_normal
    ws1.cell(row=row, column=7).fill = fill
    ws1.cell(row=row, column=7).alignment = align_right
    ws1.cell(row=row, column=7).border = thin_border
    ws1.cell(row=row, column=7).number_format = '$#,##0.00'

    # Column H: Payment Method (dropdown applied later)
    ws1.cell(row=row, column=8).font = font_normal
    ws1.cell(row=row, column=8).fill = fill
    ws1.cell(row=row, column=8).alignment = align_center
    ws1.cell(row=row, column=8).border = thin_border

    # Column I: Payment Status (dropdown applied later)
    ws1.cell(row=row, column=9).font = font_normal
    ws1.cell(row=row, column=9).fill = fill
    ws1.cell(row=row, column=9).alignment = align_center
    ws1.cell(row=row, column=9).border = thin_border

    # Column J: Date Received
    ws1.cell(row=row, column=10).font = font_normal
    ws1.cell(row=row, column=10).fill = fill
    ws1.cell(row=row, column=10).alignment = align_center
    ws1.cell(row=row, column=10).border = thin_border
    ws1.cell(row=row, column=10).number_format = "MM/DD/YYYY"

    # Column K: Notes
    ws1.cell(row=row, column=11).font = font_normal
    ws1.cell(row=row, column=11).fill = fill
    ws1.cell(row=row, column=11).alignment = align_left
    ws1.cell(row=row, column=11).border = thin_border

# Row 503: TOTAL row
total_row = 503
merge_and_style(ws1, f"A{total_row}:E{total_row}", "TOTAL", font_total, fill_total)
ws1[f"A{total_row}"].border = thin_border

cell_total_f = ws1.cell(row=total_row, column=6, value="=SUM(F3:F502)")
cell_total_f.font = font_total
cell_total_f.fill = fill_total
cell_total_f.alignment = align_right
cell_total_f.border = thin_border
cell_total_f.number_format = '$#,##0.00'

cell_total_g = ws1.cell(row=total_row, column=7, value="=SUM(G3:G502)")
cell_total_g.font = font_total
cell_total_g.fill = fill_total
cell_total_g.alignment = align_right
cell_total_g.border = thin_border
cell_total_g.number_format = '$#,##0.00'

# Fill remaining total row cells
for c in range(8, 12):
    ws1.cell(row=total_row, column=c).fill = fill_total
    ws1.cell(row=total_row, column=c).border = thin_border

# Data validations (dropdowns)
dv_service = DataValidation(
    type="list",
    formula1='"' + ",".join(SERVICES) + '"',
    allow_blank=True,
)
dv_service.error = "Please select a valid service type."
dv_service.errorTitle = "Invalid Service"
ws1.add_data_validation(dv_service)
dv_service.add("D3:D502")

dv_payment = DataValidation(
    type="list",
    formula1='"Cash,Check,Venmo,Zelle,Debit Card,ACH/Bank"',
    allow_blank=True,
)
dv_payment.error = "Please select a valid payment method."
dv_payment.errorTitle = "Invalid Payment Method"
ws1.add_data_validation(dv_payment)
dv_payment.add("H3:H502")

dv_status = DataValidation(
    type="list",
    formula1='"Paid,Partial,Pending,Overdue"',
    allow_blank=True,
)
dv_status.error = "Please select a valid payment status."
dv_status.errorTitle = "Invalid Status"
ws1.add_data_validation(dv_status)
dv_status.add("I3:I502")

# Conditional formatting on Payment Status (column I)
ws1.conditional_formatting.add(
    "I3:I502",
    CellIsRule(operator="equal", formula=['"Paid"'], fill=PatternFill("solid", fgColor=PAID_GREEN)),
)
ws1.conditional_formatting.add(
    "I3:I502",
    CellIsRule(operator="equal", formula=['"Pending"'], fill=PatternFill("solid", fgColor=PENDING_YELLOW)),
)
ws1.conditional_formatting.add(
    "I3:I502",
    CellIsRule(operator="equal", formula=['"Overdue"'], fill=PatternFill("solid", fgColor=OVERDUE_RED)),
)
ws1.conditional_formatting.add(
    "I3:I502",
    CellIsRule(operator="equal", formula=['"Partial"'], fill=PatternFill("solid", fgColor=PARTIAL_BLUE)),
)

# Freeze panes below headers
ws1.freeze_panes = "A3"

# Print setup
ws1.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws1.page_setup.orientation = "landscape"
ws1.page_setup.fitToWidth = 1
ws1.page_setup.fitToHeight = 0
ws1.print_title_rows = "1:2"

# ============================================================
# TAB 2: MONTHLY SUMMARY
# ============================================================
ws2 = wb.create_sheet("Monthly Summary")
ws2.sheet_properties.tabColor = "2D6A4F"

col_widths_2 = {1: 16, 2: 10, 3: 18, 4: 18, 5: 16, 6: 16}
for c, w in col_widths_2.items():
    ws2.column_dimensions[get_column_letter(c)].width = w

# Row 1: Title
merge_and_style(ws2, "A1:F1", "TOTALGUARD YARD CARE \u2014 2026 MONTHLY REVENUE SUMMARY", font_title, fill_title)

# Row 2: Headers
ms_headers = ["Month", "Jobs", "Revenue Charged", "Revenue Received", "Outstanding", "Collection Rate"]
for i, h in enumerate(ms_headers, 1):
    cell = ws2.cell(row=2, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Rows 3-14: Monthly data
for m in range(1, 13):
    row = m + 2
    fill = fill_white if (m - 1) % 2 == 0 else fill_gray

    # Month name
    cell_month = ws2.cell(row=row, column=1, value=MONTHS[m - 1])
    cell_month.font = font_bold
    cell_month.fill = fill
    cell_month.alignment = align_left
    cell_month.border = thin_border

    # Jobs count
    cell_jobs = ws2.cell(
        row=row, column=2,
        value=f"=SUMPRODUCT((MONTH('Revenue Log'!B$3:B$502)={m})*('Revenue Log'!F$3:F$502>0)*1)"
    )
    cell_jobs.font = font_normal
    cell_jobs.fill = fill
    cell_jobs.alignment = align_center
    cell_jobs.border = thin_border

    # Revenue Charged
    cell_charged = ws2.cell(
        row=row, column=3,
        value=f"=SUMPRODUCT((MONTH('Revenue Log'!B$3:B$502)={m})*('Revenue Log'!F$3:F$502))"
    )
    cell_charged.font = font_normal
    cell_charged.fill = fill
    cell_charged.alignment = align_right
    cell_charged.border = thin_border
    cell_charged.number_format = '$#,##0.00'

    # Revenue Received
    cell_received = ws2.cell(
        row=row, column=4,
        value=f"=SUMPRODUCT((MONTH('Revenue Log'!B$3:B$502)={m})*('Revenue Log'!G$3:G$502))"
    )
    cell_received.font = font_normal
    cell_received.fill = fill
    cell_received.alignment = align_right
    cell_received.border = thin_border
    cell_received.number_format = '$#,##0.00'

    # Outstanding
    cell_out = ws2.cell(row=row, column=5, value=f"=C{row}-D{row}")
    cell_out.font = font_normal
    cell_out.fill = fill
    cell_out.alignment = align_right
    cell_out.border = thin_border
    cell_out.number_format = '$#,##0.00'

    # Collection Rate
    cell_rate = ws2.cell(row=row, column=6, value=f"=IF(C{row}>0,D{row}/C{row},0)")
    cell_rate.font = font_normal
    cell_rate.fill = fill
    cell_rate.alignment = align_center
    cell_rate.border = thin_border
    cell_rate.number_format = '0.0%'

# Row 15: Annual total
total_row_ms = 15
ws2.cell(row=total_row_ms, column=1, value="ANNUAL TOTAL").font = font_total
ws2.cell(row=total_row_ms, column=1).fill = fill_total
ws2.cell(row=total_row_ms, column=1).alignment = align_left
ws2.cell(row=total_row_ms, column=1).border = thin_border

for col in range(2, 6):
    cell = ws2.cell(row=total_row_ms, column=col, value=f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}14)")
    cell.font = font_total
    cell.fill = fill_total
    cell.alignment = align_right if col >= 3 else align_center
    cell.border = thin_border
    if col >= 3:
        cell.number_format = '$#,##0.00'

# Collection rate for annual total
cell_annual_rate = ws2.cell(row=total_row_ms, column=6, value="=IF(C15>0,D15/C15,0)")
cell_annual_rate.font = font_total
cell_annual_rate.fill = fill_total
cell_annual_rate.alignment = align_center
cell_annual_rate.border = thin_border
cell_annual_rate.number_format = '0.0%'

# Row 17: Average Job Size
ws2.cell(row=17, column=1, value="Average Job Size").font = font_bold
ws2.cell(row=17, column=1).alignment = align_left
cell_avg = ws2.cell(row=17, column=2, value="=IF(B15>0,C15/B15,0)")
cell_avg.font = font_total
cell_avg.alignment = align_right
cell_avg.number_format = '$#,##0.00'

# Freeze panes
ws2.freeze_panes = "A3"

# Print setup
ws2.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws2.page_setup.orientation = "landscape"
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 0
ws2.print_title_rows = "1:2"

# ============================================================
# TAB 3: REVENUE BY SERVICE
# ============================================================
ws3 = wb.create_sheet("Revenue by Service")
ws3.sheet_properties.tabColor = "52B788"

col_widths_3 = {1: 26, 2: 10, 3: 18, 4: 18, 5: 14}
for c, w in col_widths_3.items():
    ws3.column_dimensions[get_column_letter(c)].width = w

# Row 1: Title
merge_and_style(ws3, "A1:E1", "TOTALGUARD YARD CARE \u2014 2026 REVENUE BY SERVICE TYPE", font_title, fill_title)

# Row 2: Headers
svc_headers = ["Service", "Jobs", "Total Revenue", "Avg Revenue/Job", "% of Total"]
for i, h in enumerate(svc_headers, 1):
    cell = ws3.cell(row=2, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Total row number (after 15 services + header rows)
svc_total_row = 3 + len(SERVICES)  # row 18

# Rows 3-17: Service rows
for idx, svc in enumerate(SERVICES):
    row = 3 + idx
    fill = fill_white if idx % 2 == 0 else fill_gray

    # Service name
    cell_svc = ws3.cell(row=row, column=1, value=svc)
    cell_svc.font = font_bold
    cell_svc.fill = fill
    cell_svc.alignment = align_left
    cell_svc.border = thin_border

    # Jobs count
    cell_jobs = ws3.cell(
        row=row, column=2,
        value=f'=COUNTIF(\'Revenue Log\'!D$3:D$502,A{row})'
    )
    cell_jobs.font = font_normal
    cell_jobs.fill = fill
    cell_jobs.alignment = align_center
    cell_jobs.border = thin_border

    # Total Revenue
    cell_rev = ws3.cell(
        row=row, column=3,
        value=f'=SUMIF(\'Revenue Log\'!D$3:D$502,A{row},\'Revenue Log\'!F$3:F$502)'
    )
    cell_rev.font = font_normal
    cell_rev.fill = fill
    cell_rev.alignment = align_right
    cell_rev.border = thin_border
    cell_rev.number_format = '$#,##0.00'

    # Avg Revenue/Job
    cell_avg = ws3.cell(row=row, column=4, value=f"=IF(B{row}>0,C{row}/B{row},0)")
    cell_avg.font = font_normal
    cell_avg.fill = fill
    cell_avg.alignment = align_right
    cell_avg.border = thin_border
    cell_avg.number_format = '$#,##0.00'

    # % of Total
    cell_pct = ws3.cell(row=row, column=5, value=f"=IF(C${svc_total_row}>0,C{row}/C${svc_total_row},0)")
    cell_pct.font = font_normal
    cell_pct.fill = fill
    cell_pct.alignment = align_center
    cell_pct.border = thin_border
    cell_pct.number_format = '0.0%'

# Total row
ws3.cell(row=svc_total_row, column=1, value="TOTAL").font = font_total
ws3.cell(row=svc_total_row, column=1).fill = fill_total
ws3.cell(row=svc_total_row, column=1).alignment = align_left
ws3.cell(row=svc_total_row, column=1).border = thin_border

cell_total_jobs = ws3.cell(row=svc_total_row, column=2, value=f"=SUM(B3:B{svc_total_row - 1})")
cell_total_jobs.font = font_total
cell_total_jobs.fill = fill_total
cell_total_jobs.alignment = align_center
cell_total_jobs.border = thin_border

cell_total_rev = ws3.cell(row=svc_total_row, column=3, value=f"=SUM(C3:C{svc_total_row - 1})")
cell_total_rev.font = font_total
cell_total_rev.fill = fill_total
cell_total_rev.alignment = align_right
cell_total_rev.border = thin_border
cell_total_rev.number_format = '$#,##0.00'

cell_total_avg = ws3.cell(row=svc_total_row, column=4, value=f"=IF(B{svc_total_row}>0,C{svc_total_row}/B{svc_total_row},0)")
cell_total_avg.font = font_total
cell_total_avg.fill = fill_total
cell_total_avg.alignment = align_right
cell_total_avg.border = thin_border
cell_total_avg.number_format = '$#,##0.00'

cell_total_pct = ws3.cell(row=svc_total_row, column=5, value="100.0%")
cell_total_pct.font = font_total
cell_total_pct.fill = fill_total
cell_total_pct.alignment = align_center
cell_total_pct.border = thin_border

# Freeze panes
ws3.freeze_panes = "A3"

# Print setup
ws3.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws3.page_setup.orientation = "landscape"
ws3.page_setup.fitToWidth = 1
ws3.page_setup.fitToHeight = 0
ws3.print_title_rows = "1:2"

# ============================================================
# SAVE
# ============================================================
output_path = r"c:\Users\vance\OneDrive\Desktop\claude-workspace\TotalGuard_2026_Revenue_Tracker.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
