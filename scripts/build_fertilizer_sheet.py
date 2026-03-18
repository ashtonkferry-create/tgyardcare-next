import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()

# ============================================================
# COLOR PALETTE
# ============================================================
DARK_GREEN = "1B4332"
MED_GREEN = "2D6A4F"
LIGHT_GREEN = "40916C"
ACCENT_GREEN = "52B788"
DARK_BG = "1A1A2E"
HEADER_BG = "1B4332"
SUB_HEADER_BG = "2D6A4F"
INPUT_BG = "FFFFF0"
INPUT_BORDER = "FF8C00"
ROW_ALT_1 = "F0F7F4"
ROW_ALT_2 = "FFFFFF"
PROFIT_GREEN = "D4EDDA"
COST_RED = "F8D7DA"
TOTAL_BG = "D4EDDA"
LINK_BLUE = "0563C1"
WHITE = "FFFFFF"
BLACK = "000000"
GOLD = "FFD700"

# Reusable styles
font_title = Font(name="Arial", size=20, bold=True, color=WHITE)
font_subtitle = Font(name="Arial", size=12, italic=True, color=WHITE)
font_header = Font(name="Arial", size=11, bold=True, color=WHITE)
font_sub_header = Font(name="Arial", size=10, bold=True, color=WHITE)
font_normal = Font(name="Arial", size=10, color=BLACK)
font_bold = Font(name="Arial", size=10, bold=True, color=BLACK)
font_input = Font(name="Arial", size=14, bold=True, color="0000FF")
font_input_label = Font(name="Arial", size=12, bold=True, color=DARK_GREEN)
font_total = Font(name="Arial", size=13, bold=True, color=DARK_GREEN)
font_big_total = Font(name="Arial", size=16, bold=True, color=WHITE)
font_link = Font(name="Arial", size=10, color=LINK_BLUE, underline="single")
font_step = Font(name="Arial", size=12, bold=True, color=DARK_GREEN)
font_gold = Font(name="Arial", size=11, bold=True, color="8B6914")

fill_header = PatternFill("solid", fgColor=HEADER_BG)
fill_sub_header = PatternFill("solid", fgColor=SUB_HEADER_BG)
fill_input = PatternFill("solid", fgColor=INPUT_BG)
fill_alt1 = PatternFill("solid", fgColor=ROW_ALT_1)
fill_alt2 = PatternFill("solid", fgColor=ROW_ALT_2)
fill_total = PatternFill("solid", fgColor=TOTAL_BG)
fill_dark = PatternFill("solid", fgColor=DARK_BG)
fill_profit = PatternFill("solid", fgColor=PROFIT_GREEN)
fill_cost = PatternFill("solid", fgColor=COST_RED)
fill_gold = PatternFill("solid", fgColor="FFF8DC")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="center")

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
thick_border = Border(
    left=Side(style="medium", color=DARK_GREEN),
    right=Side(style="medium", color=DARK_GREEN),
    top=Side(style="medium", color=DARK_GREEN),
    bottom=Side(style="medium", color=DARK_GREEN),
)
input_border = Border(
    left=Side(style="thick", color=INPUT_BORDER),
    right=Side(style="thick", color=INPUT_BORDER),
    top=Side(style="thick", color=INPUT_BORDER),
    bottom=Side(style="thick", color=INPUT_BORDER),
)

def style_range(ws, row, cols, font=None, fill=None, alignment=None, border=None, number_format=None):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if number_format: cell.number_format = number_format

def merge_and_style(ws, range_str, value, font, fill, alignment=align_center):
    ws.merge_cells(range_str)
    top_left = range_str.split(":")[0]
    cell = ws[top_left]
    cell.value = value
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment

# ============================================================
# SHEET 1: QUOTE CALCULATOR
# ============================================================
ws1 = wb.active
ws1.title = "QUOTE CALCULATOR"
ws1.sheet_properties.tabColor = DARK_GREEN

# Column widths
col_widths_1 = {1: 4, 2: 30, 3: 22, 4: 16, 5: 16, 6: 18, 7: 18, 8: 4}
for c, w in col_widths_1.items():
    ws1.column_dimensions[get_column_letter(c)].width = w

# Row 1-2: Title banner
merge_and_style(ws1, "A1:H2", "TOTALGUARD YARD CARE — FERTILIZER MASTER SHEET", font_title, fill_header)

# Row 3: Subtitle
merge_and_style(ws1, "A3:H3", "5-Step Professional Lawn Program  |  Wisconsin Cool-Season Grasses  |  SiteOne Products", font_subtitle, fill_sub_header)

# Row 5-6: Input section
ws1.merge_cells("B5:C5")
ws1["B5"].value = "ENTER PROPERTY SQUARE FOOTAGE:"
ws1["B5"].font = font_input_label
ws1["B5"].alignment = Alignment(horizontal="right", vertical="center")

ws1["D5"].value = 10000
ws1["D5"].font = font_input
ws1["D5"].fill = fill_input
ws1["D5"].border = input_border
ws1["D5"].alignment = align_center
ws1["D5"].number_format = "#,##0"

ws1.merge_cells("E5:F5")
ws1["E5"].value = "sqft"
ws1["E5"].font = Font(name="Arial", size=12, color="888888")
ws1["E5"].alignment = Alignment(horizontal="left", vertical="center")

# Row 6: Markup multiplier
ws1.merge_cells("B6:C6")
ws1["B6"].value = "YOUR MARKUP MULTIPLIER:"
ws1["B6"].font = font_input_label
ws1["B6"].alignment = Alignment(horizontal="right", vertical="center")

ws1["D6"].value = 2.5
ws1["D6"].font = font_input
ws1["D6"].fill = fill_input
ws1["D6"].border = input_border
ws1["D6"].alignment = align_center
ws1["D6"].number_format = "0.0x"

ws1.merge_cells("E6:F6")
ws1["E6"].value = "(2.0 = double, 2.5 = standard, 3.0 = premium)"
ws1["E6"].font = Font(name="Arial", size=9, italic=True, color="888888")
ws1["E6"].alignment = Alignment(horizontal="left", vertical="center")

# Row 8: Column headers for calculator
r = 8
headers = ["", "STEP", "PRODUCT", "BAGS NEEDED", "MATERIAL COST", "YOUR PRICE", "PROFIT", ""]
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=r, column=i, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = thin_border

# Product data references (will pull from PRODUCT DATABASE sheet)
# Format: [step_name, product_name, coverage_per_bag, price_per_bag]
steps = [
    ["Step 1: Pre-Emergent + Fert", "LESCO Stonewall 0.43% 19-0-7", 12500, 28],
    ["Step 2: Weed Control + Feed", "LESCO 24-0-6 + Trimec Classic", 12000, 27],
    ["Step 3: Grub Prevention + Fert", "LESCO Merit 0.2% 14-0-0", 14200, 65],
    ["Step 4: Fall Recovery", "LESCO 24-0-11 50% PolyPlus", 12000, 48],
    ["Step 5: Winterizer", "LESCO 32-0-10 100% NOS Plus SOP", 16000, 63],
]

# Step 2 also needs Trimec - we'll handle the herbicide cost separately
# Trimec Classic: ~$60/gal covers ~85,000 sqft at 1.5oz/1000sqft
# Step 4 also needs Trimec for fall weed control

for i, (step_name, product, coverage, price) in enumerate(steps):
    r = 9 + i
    fill = fill_alt1 if i % 2 == 0 else fill_alt2

    ws1.cell(row=r, column=1).fill = fill

    cell_step = ws1.cell(row=r, column=2, value=step_name)
    cell_step.font = font_step
    cell_step.fill = fill
    cell_step.alignment = align_left
    cell_step.border = thin_border

    cell_prod = ws1.cell(row=r, column=3, value=product)
    cell_prod.font = font_normal
    cell_prod.fill = fill
    cell_prod.alignment = align_left
    cell_prod.border = thin_border

    # Bags needed = ROUNDUP(sqft / coverage, 0)
    bags_formula = f'=ROUNDUP($D$5/{coverage},0)'
    cell_bags = ws1.cell(row=r, column=4, value=bags_formula)
    cell_bags.font = font_bold
    cell_bags.fill = fill
    cell_bags.alignment = align_center
    cell_bags.border = thin_border
    cell_bags.number_format = "0"

    # Material cost = bags * price_per_bag
    # For steps 2 and 4, add Trimec cost
    if i == 1:  # Step 2 - add Trimec
        cost_formula = f'=D{r}*{price}+ROUNDUP($D$5/85000,0)*60'
    elif i == 3:  # Step 4 - add Trimec
        cost_formula = f'=D{r}*{price}+ROUNDUP($D$5/85000,0)*60'
    else:
        cost_formula = f'=D{r}*{price}'

    cell_cost = ws1.cell(row=r, column=5, value=cost_formula)
    cell_cost.font = font_bold
    cell_cost.fill = fill
    cell_cost.alignment = align_center
    cell_cost.border = thin_border
    cell_cost.number_format = "$#,##0.00"

    # Your price = material cost * markup
    price_formula = f'=E{r}*$D$6'
    cell_price = ws1.cell(row=r, column=6, value=price_formula)
    cell_price.font = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
    cell_price.fill = fill
    cell_price.alignment = align_center
    cell_price.border = thin_border
    cell_price.number_format = "$#,##0.00"

    # Profit = your price - material cost
    profit_formula = f'=F{r}-E{r}'
    cell_profit = ws1.cell(row=r, column=7, value=profit_formula)
    cell_profit.font = Font(name="Arial", size=10, bold=True, color="228B22")
    cell_profit.fill = fill
    cell_profit.alignment = align_center
    cell_profit.border = thin_border
    cell_profit.number_format = "$#,##0.00"

    ws1.cell(row=r, column=8).fill = fill

# Row 14: Totals row
r_total = 14
merge_and_style(ws1, f"B{r_total}:C{r_total}", "FULL 5-STEP PROGRAM TOTALS", font_big_total, PatternFill("solid", fgColor=DARK_GREEN))
ws1[f"B{r_total}"].border = thick_border
ws1[f"C{r_total}"].border = thick_border

ws1.cell(row=r_total, column=4).fill = PatternFill("solid", fgColor=DARK_GREEN)
ws1.cell(row=r_total, column=4).border = thick_border

cell_total_cost = ws1.cell(row=r_total, column=5, value="=SUM(E9:E13)")
cell_total_cost.font = Font(name="Arial", size=14, bold=True, color=WHITE)
cell_total_cost.fill = PatternFill("solid", fgColor=DARK_GREEN)
cell_total_cost.alignment = align_center
cell_total_cost.border = thick_border
cell_total_cost.number_format = "$#,##0.00"

cell_total_price = ws1.cell(row=r_total, column=6, value="=SUM(F9:F13)")
cell_total_price.font = Font(name="Arial", size=14, bold=True, color=GOLD)
cell_total_price.fill = PatternFill("solid", fgColor=DARK_GREEN)
cell_total_price.alignment = align_center
cell_total_price.border = thick_border
cell_total_price.number_format = "$#,##0.00"

cell_total_profit = ws1.cell(row=r_total, column=7, value="=SUM(G9:G13)")
cell_total_profit.font = Font(name="Arial", size=14, bold=True, color=GOLD)
cell_total_profit.fill = PatternFill("solid", fgColor=DARK_GREEN)
cell_total_profit.alignment = align_center
cell_total_profit.border = thick_border
cell_total_profit.number_format = "$#,##0.00"

ws1.cell(row=r_total, column=8).fill = PatternFill("solid", fgColor=DARK_GREEN)

# Row 16: Per-visit cost
ws1.merge_cells("B16:C16")
ws1["B16"].value = "MATERIAL COST PER VISIT (avg):"
ws1["B16"].font = font_bold
ws1["B16"].alignment = Alignment(horizontal="right", vertical="center")
ws1["E16"].value = "=E14/5"
ws1["E16"].font = font_bold
ws1["E16"].number_format = "$#,##0.00"
ws1["E16"].alignment = align_center

ws1.merge_cells("B17:C17")
ws1["B17"].value = "YOUR CHARGE PER VISIT (avg):"
ws1["B17"].font = font_bold
ws1["B17"].alignment = Alignment(horizontal="right", vertical="center")
ws1["F17"].value = "=F14/5"
ws1["F17"].font = Font(name="Arial", size=11, bold=True, color=DARK_GREEN)
ws1["F17"].number_format = "$#,##0.00"
ws1["F17"].alignment = align_center

# Row 19: Quick reference header
r_qr = 19
merge_and_style(ws1, f"B{r_qr}:G{r_qr}", "QUICK REFERENCE — COMMON PROPERTY SIZES", font_header, fill_header)

# Row 20: QR column headers
r_qrh = 20
qr_headers = ["", "PROPERTY SIZE", "TOTAL MATERIAL", "YOUR CHARGE", "YOUR PROFIT", "PROFIT MARGIN", ""]
for i, h in enumerate(qr_headers, 1):
    cell = ws1.cell(row=r_qrh, column=i+1 if i > 1 else 1, value=h)
    if i >= 2:
        ci = i
        cell = ws1.cell(row=r_qrh, column=ci, value=h)
        cell.font = font_sub_header
        cell.fill = fill_sub_header
        cell.alignment = align_center
        cell.border = thin_border

# Quick reference sizes
qr_sizes = [5000, 8000, 10000, 15000, 20000, 40000, 80000]
for idx, size in enumerate(qr_sizes):
    r = r_qrh + 1 + idx
    fill = fill_alt1 if idx % 2 == 0 else fill_alt2

    ws1.cell(row=r, column=2, value=f"{size:,} sqft")
    ws1.cell(row=r, column=2).font = font_bold
    ws1.cell(row=r, column=2).fill = fill
    ws1.cell(row=r, column=2).alignment = align_center
    ws1.cell(row=r, column=2).border = thin_border

    # Total material cost for this size
    # Sum of: ROUNDUP(size/coverage)*price for each step + Trimec for steps 2&4
    mat_formula = (
        f'=ROUNDUP({size}/12500,0)*28'  # Step 1
        f'+ROUNDUP({size}/12000,0)*27+ROUNDUP({size}/85000,0)*60'  # Step 2 + Trimec
        f'+ROUNDUP({size}/14200,0)*65'  # Step 3
        f'+ROUNDUP({size}/12000,0)*48+ROUNDUP({size}/85000,0)*60'  # Step 4 + Trimec
        f'+ROUNDUP({size}/16000,0)*63'  # Step 5
    )
    cell_mat = ws1.cell(row=r, column=3, value=mat_formula)
    cell_mat.font = font_bold
    cell_mat.fill = fill
    cell_mat.alignment = align_center
    cell_mat.border = thin_border
    cell_mat.number_format = "$#,##0"

    # Your charge
    charge_formula = f'=C{r}*$D$6'
    cell_charge = ws1.cell(row=r, column=4, value=charge_formula)
    cell_charge.font = Font(name="Arial", size=10, bold=True, color=DARK_GREEN)
    cell_charge.fill = fill
    cell_charge.alignment = align_center
    cell_charge.border = thin_border
    cell_charge.number_format = "$#,##0"

    # Profit
    profit_formula = f'=D{r}-C{r}'
    cell_prof = ws1.cell(row=r, column=5, value=profit_formula)
    cell_prof.font = Font(name="Arial", size=10, bold=True, color="228B22")
    cell_prof.fill = fill
    cell_prof.alignment = align_center
    cell_prof.border = thin_border
    cell_prof.number_format = "$#,##0"

    # Margin
    margin_formula = f'=E{r}/D{r}'
    cell_margin = ws1.cell(row=r, column=6, value=margin_formula)
    cell_margin.font = font_bold
    cell_margin.fill = fill
    cell_margin.alignment = align_center
    cell_margin.border = thin_border
    cell_margin.number_format = "0.0%"

# Row after QR table: Notes
r_notes = r_qrh + 1 + len(qr_sizes) + 1
ws1.merge_cells(f"B{r_notes}:G{r_notes}")
ws1[f"B{r_notes}"].value = "NOTE: Prices based on SiteOne estimated pricing. Actual costs may vary by branch/volume. Trimec herbicide included in Steps 2 & 4."
ws1[f"B{r_notes}"].font = Font(name="Arial", size=9, italic=True, color="666666")
ws1[f"B{r_notes}"].alignment = align_left

ws1.merge_cells(f"B{r_notes+1}:G{r_notes+1}")
ws1[f"B{r_notes+1}"].value = "TIP: Change the blue input cells (sq footage & markup) to recalculate everything instantly."
ws1[f"B{r_notes+1}"].font = Font(name="Arial", size=9, italic=True, color="0000FF")
ws1[f"B{r_notes+1}"].alignment = align_left

# Freeze panes
ws1.freeze_panes = "A8"

# Print setup
ws1.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

# ============================================================
# SHEET 2: PRODUCT DATABASE
# ============================================================
ws2 = wb.create_sheet("PRODUCT DATABASE")
ws2.sheet_properties.tabColor = MED_GREEN

col_widths_2 = {1: 3, 2: 14, 3: 38, 4: 12, 5: 14, 6: 16, 7: 14, 8: 20, 9: 55, 10: 3}
for c, w in col_widths_2.items():
    ws2.column_dimensions[get_column_letter(c)].width = w

# Title
merge_and_style(ws2, "A1:I2", "PRODUCT DATABASE — SITEONE LANDSCAPE SUPPLY", font_title, fill_header)
merge_and_style(ws2, "A3:I3", "All products verified on SiteOne.com  |  Prices are estimates (login required for exact pricing)  |  Links are clickable", font_subtitle, fill_sub_header)

# Product database entries
products = [
    {
        "step": "STEP 1 — EARLY SPRING: Pre-Emergent + Fertilizer (Late April – Early May)",
        "timing": "Apply when soil temps reach 55°F consistently",
        "items": [
            {
                "type": "RECOMMENDED",
                "name": "LESCO Stonewall 0.43% 19-0-7 30% PolyPlus OPTI45",
                "sku": "702053",
                "npk": "19-0-7",
                "bag_size": "50 lb",
                "coverage": "12,500 sqft",
                "price": "$23–$35",
                "notes": "Best value. LESCO house brand = cheapest at SiteOne. 0.43% Prodiamine is strong crabgrass barrier.",
                "url": "https://www.siteone.com/en/702053-lesco-stonewall-043-19-0-7-30-polyplus-opti45-mop-pre-emergent-granular-herbicide-plus-fertilizer-50-lb-bag/p/356351"
            },
            {
                "type": "ALTERNATIVE",
                "name": "Lebanon Pro Prodiamine 0.38% 18-0-4 25% CRN",
                "sku": "2241024",
                "npk": "18-0-4",
                "bag_size": "50 lb",
                "coverage": "12,500 sqft",
                "price": "$38–$45",
                "notes": "Better NPK for spring feeding. Higher price but premium quality.",
                "url": "https://www.siteone.com/en/2241024-lebanon-pro-prodiamine-038-18-0-4-25-crn-pre-emergent-herbicide-plus-fertilizer-50-lb/p/780809"
            },
            {
                "type": "PREMIUM",
                "name": "LESCO Dimension 0.15% 19-0-0 30% PolyPlus OPTI45",
                "sku": "702022",
                "npk": "19-0-0",
                "bag_size": "50 lb",
                "coverage": "12,500 sqft",
                "price": "$25–$35",
                "notes": "Dimension (Dithiopyr) has early post-emergent activity on young crabgrass that Prodiamine does not.",
                "url": "https://www.siteone.com/en/702022-lesco-dimension-015-19-0-0-30-polyplus-opti45-pre-emergent-plus-fertilizer-50-lb/p/356322"
            },
        ]
    },
    {
        "step": "STEP 2 — LATE SPRING: Weed Control + Feeding (Late May – Early June)",
        "timing": "Spray when weeds are actively growing. Spread fertilizer same week.",
        "items": [
            {
                "type": "RECOMMENDED (Fertilizer)",
                "name": "LESCO 24-0-6 50% PolyPlus OPTI BIO MOP",
                "sku": "511484",
                "npk": "24-0-6",
                "bag_size": "50 lb",
                "coverage": "12,000 sqft",
                "price": "$19–$35",
                "notes": "50% controlled-release nitrogen. Biosolids for organic matter. Great balanced feed.",
                "url": "https://www.siteone.com/en/511484-lesco-24-0-6-50-polyplus-opti-granular-fertilizer-50-lb-bag/p/852704"
            },
            {
                "type": "RECOMMENDED (Herbicide)",
                "name": "Trimec Classic Broadleaf Herbicide — 1 gal",
                "sku": "8811076",
                "npk": "N/A",
                "bag_size": "1 gallon",
                "coverage": "~85,000 sqft",
                "price": "$55–$70",
                "notes": "Industry standard 3-way (2,4-D + Dicamba + MCPP). Controls 90+ broadleaf weeds. 1.5 oz/1,000 sqft rate.",
                "url": "https://www.siteone.com/en/8811076-trimec-classic-broadleaf-post-emergent-liquid-herbicide-1-gal/p/163544"
            },
            {
                "type": "BULK (Herbicide)",
                "name": "Trimec Classic Broadleaf Herbicide — 2.5 gal",
                "sku": "8811126",
                "npk": "N/A",
                "bag_size": "2.5 gallon",
                "coverage": "~213,000 sqft",
                "price": "$130–$170",
                "notes": "Better value for high volume. Same product, bigger jug. Requires pesticide applicator license.",
                "url": "https://www.siteone.com/en/8811126-trimec-classic-broadleaf-post-emergent-liquid-herbicide-25-gal/p/163545"
            },
            {
                "type": "ALTERNATIVE (Herbicide)",
                "name": "Nufarm Triplet SF Herbicide — 2.5 gal",
                "sku": "10447225",
                "npk": "N/A",
                "bag_size": "2.5 gallon",
                "coverage": "~213,000 sqft",
                "price": "$80–$110",
                "notes": "Generic 3-way blend. Cheaper than Trimec, same active ingredients. Good budget option.",
                "url": "https://www.siteone.com/en/10447225-nufarm-triplet-sf-herbicide-25-gal/p/772634"
            },
        ]
    },
    {
        "step": "STEP 3 — EARLY SUMMER: Grub Prevention + Fertilizer (Late June – Early July)",
        "timing": "Apply before grub eggs hatch. Preventative, not curative.",
        "items": [
            {
                "type": "RECOMMENDED (Combo)",
                "name": "LESCO Merit 0.2% 14-0-0 30% PolyPlus OPTI45 Insecticide + Fert",
                "sku": "291485",
                "npk": "14-0-0",
                "bag_size": "50 lb",
                "coverage": "14,200 sqft",
                "price": "$55–$75",
                "notes": "One-pass combo: grub prevention + summer feeding. Merit (Imidacloprid) is the industry standard.",
                "url": "https://www.siteone.com/en/291485-lesco-merit-02-14-0-0-30polyplus-opti45-systemic-granular-insecticide-plus-fertilizer-50-lb-bag/p/337866"
            },
            {
                "type": "PREMIUM (Combo)",
                "name": "LESCO Acelepryn 0.067% 20-0-5 75% PolyPlus OPTI45 2Fe",
                "sku": "291610",
                "npk": "20-0-5",
                "bag_size": "50 lb",
                "coverage": "14,500 sqft",
                "price": "$85–$110",
                "notes": "Premium option. Acelepryn (Chlorantraniliprole) = lower pollinator toxicity, longer residual. Better NPK too.",
                "url": "https://www.siteone.com/en/291610-lesco-acelepryn-0067-20-0-5-75-polyplus-opti45-2fe-insecticide-plus-fertilizer-50-lb/p/364577"
            },
            {
                "type": "BUDGET (Separate)",
                "name": "Quali-Pro Imidacloprid 0.5G (grub only) + LESCO 18-0-3",
                "sku": "83014299",
                "npk": "N/A + 18-0-3",
                "bag_size": "30 lb + 50 lb",
                "coverage": "16,500 + 14,000 sqft",
                "price": "$30–$45 + $35–$50",
                "notes": "Generic imidacloprid for grubs + separate light summer fert. Two passes but cheapest option.",
                "url": "https://www.siteone.com/en/83014299-control-solutions-quali-pro-imidacloprid-05g-insecticide/p/772586"
            },
        ]
    },
    {
        "step": "STEP 4 — LATE SUMMER/EARLY FALL: Lawn Recovery (Late Aug – Early Sep)",
        "timing": "Repair summer stress. This is when grass starts growing aggressively again.",
        "items": [
            {
                "type": "RECOMMENDED",
                "name": "LESCO 24-0-11 50% PolyPlus OPTI 2% Fe 1% Mn MOP",
                "sku": "098631",
                "npk": "24-0-11",
                "bag_size": "50 lb",
                "coverage": "12,000 sqft",
                "price": "$40–$55",
                "notes": "Iron + Manganese = dark green color. Potassium builds winter hardiness. The recovery king.",
                "url": "https://www.siteone.com/en/098631-lesco-24-0-11-50-polyplus-opti-2fe-1mn-mop-turfgrass-granular-fertilizer-50-lb-bag/p/336709"
            },
            {
                "type": "ALTERNATIVE",
                "name": "LESCO 24-0-10 75% PolyPlus OPTI45 Spar-TECH",
                "sku": "511673",
                "npk": "24-0-10",
                "bag_size": "50 lb",
                "coverage": "12,000 sqft",
                "price": "$45–$60",
                "notes": "75% slow-release = feeds longer. Spar-TECH coating technology. Premium option.",
                "url": "https://www.siteone.com/en/511673-lesco-24-0-10-75-polyplus-opti45-spar-tech-10-cl-mop-turfgrass-granular-fertilizer-50-lb-bag/p/981609"
            },
            {
                "type": "HERBICIDE (pair with above)",
                "name": "Trimec Classic — same as Step 2",
                "sku": "8811076",
                "npk": "N/A",
                "bag_size": "1 gallon",
                "coverage": "~85,000 sqft",
                "price": "$55–$70",
                "notes": "Fall is the BEST time to spray broadleaf weeds in Wisconsin. Weeds are pulling nutrients to roots = herbicide goes deeper.",
                "url": "https://www.siteone.com/en/8811076-trimec-classic-broadleaf-post-emergent-liquid-herbicide-1-gal/p/163544"
            },
        ]
    },
    {
        "step": "STEP 5 — LATE FALL: Winterizer (Mid October – Early November)",
        "timing": "Last application of the year. Grass is still growing roots even after top growth stops.",
        "items": [
            {
                "type": "RECOMMENDED",
                "name": "LESCO 32-0-10 100% NOS Plus SOP Mini Granular",
                "sku": "511153",
                "npk": "32-0-10",
                "bag_size": "50 lb",
                "coverage": "16,000 sqft",
                "price": "$55–$75",
                "notes": "THE winterizer. NOS Plus stabilized nitrogen maximizes root storage. SOP = chloride-free potassium. Mini granule = best coverage.",
                "url": "https://www.siteone.com/en/511153-lesco-32-0-10-100-nos-plus-sop-turfgrass-mini-granular-fertilizer-50-lb-bag/p/682926"
            },
            {
                "type": "ALTERNATIVE",
                "name": "LESCO 32-0-10 75% PolyPlus 2% Fe",
                "sku": "510124",
                "npk": "32-0-10",
                "bag_size": "50 lb",
                "coverage": "16,000 sqft",
                "price": "$50–$70",
                "notes": "Same NPK, polymer-coated instead of NOS. 2% Iron for late-season color. Slightly cheaper.",
                "url": "https://www.siteone.com/en/510124-lesco-32-0-10-75-polyplus-2-fe-turfgrass-granular-fertilizer-50-lb-bag/p/351438"
            },
            {
                "type": "BUDGET",
                "name": "LESCO 30-0-10 25% PolyPlus",
                "sku": "510120",
                "npk": "30-0-10",
                "bag_size": "50 lb",
                "coverage": "18,750 sqft",
                "price": "$45–$65",
                "notes": "Covers more area per bag (18,750 sqft). Lower slow-release %. Good budget winterizer.",
                "url": "https://www.siteone.com/en/510120-lesco-30-0-10-25-polyplus-turfgrass-granular-fertilizer-50-lb-bag/p/351435"
            },
        ]
    },
]

r = 5
for step_data in products:
    # Step header
    merge_and_style(ws2, f"B{r}:I{r}", step_data["step"], Font(name="Arial", size=13, bold=True, color=WHITE), fill_header)
    r += 1

    # Timing note
    merge_and_style(ws2, f"B{r}:I{r}", step_data["timing"], Font(name="Arial", size=10, italic=True, color=WHITE), fill_sub_header)
    r += 1

    # Column headers for products
    prod_headers = ["Type", "Product Name", "SKU", "NPK", "Bag Size", "Coverage", "Est. Price", "Notes / Why"]
    for i, h in enumerate(prod_headers):
        cell = ws2.cell(row=r, column=i+2, value=h)
        cell.font = Font(name="Arial", size=9, bold=True, color=DARK_GREEN)
        cell.fill = PatternFill("solid", fgColor="E8F5E9")
        cell.alignment = align_center
        cell.border = thin_border
    r += 1

    # Product rows
    for idx, item in enumerate(step_data["items"]):
        fill = fill_alt1 if idx % 2 == 0 else fill_alt2
        if "RECOMMENDED" in item["type"]:
            fill = fill_gold

        ws2.cell(row=r, column=2, value=item["type"]).font = Font(name="Arial", size=9, bold=True, color="8B6914" if "RECOMMENDED" in item["type"] else BLACK)
        ws2.cell(row=r, column=2).fill = fill
        ws2.cell(row=r, column=2).alignment = align_center
        ws2.cell(row=r, column=2).border = thin_border

        # Product name as hyperlink
        cell_name = ws2.cell(row=r, column=3, value=item["name"])
        cell_name.hyperlink = item["url"]
        cell_name.font = font_link
        cell_name.fill = fill
        cell_name.alignment = align_left
        cell_name.border = thin_border

        for col_idx, key in enumerate(["sku", "npk", "bag_size", "coverage", "price"], 4):
            cell = ws2.cell(row=r, column=col_idx, value=item[key])
            cell.font = font_normal
            cell.fill = fill
            cell.alignment = align_center
            cell.border = thin_border

        ws2.cell(row=r, column=9, value=item["notes"]).font = Font(name="Arial", size=9, color="555555")
        ws2.cell(row=r, column=9).fill = fill
        ws2.cell(row=r, column=9).alignment = align_left
        ws2.cell(row=r, column=9).border = thin_border

        r += 1

    r += 1  # Blank row between steps

# Add final notes
r += 1
merge_and_style(ws2, f"B{r}:I{r}", "IMPORTANT NOTES", Font(name="Arial", size=12, bold=True, color=WHITE), fill_header)
r += 1
notes = [
    "All prices are ESTIMATES. SiteOne requires a professional account login for exact pricing. Call your local branch.",
    "LESCO is SiteOne's house brand — always the best pricing at SiteOne locations.",
    "Trimec Classic requires a pesticide applicator license to purchase.",
    "Wisconsin SiteOne locations: Madison, Milwaukee, Green Bay, Appleton, Waukesha, Brookfield.",
    "Buy by the pallet for best pricing. Typical pro discount is 15-25% below retail.",
    "Product links are clickable — they go directly to the SiteOne product page.",
]
for note in notes:
    merge_and_style(ws2, f"B{r}:I{r}", f"• {note}", Font(name="Arial", size=9, color="444444"), fill_alt1, align_left)
    r += 1

ws2.freeze_panes = "A5"

# ============================================================
# SHEET 3: PROFIT ANALYSIS
# ============================================================
ws3 = wb.create_sheet("PROFIT ANALYSIS")
ws3.sheet_properties.tabColor = ACCENT_GREEN

col_widths_3 = {1: 3, 2: 30, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16, 8: 16, 9: 3}
for c, w in col_widths_3.items():
    ws3.column_dimensions[get_column_letter(c)].width = w

# Title
merge_and_style(ws3, "A1:H2", "PROFIT ANALYSIS — FERTILIZER PROGRAM", font_title, fill_header)
merge_and_style(ws3, "A3:H3", "Material cost vs. revenue at different markup levels  |  Linked to Quote Calculator inputs", font_subtitle, fill_sub_header)

# Section 1: Markup comparison
r = 5
merge_and_style(ws3, f"B{r}:H{r}", "MARKUP COMPARISON (based on Quote Calculator sq footage)", Font(name="Arial", size=12, bold=True, color=WHITE), fill_header)
r += 1

# Headers
markup_headers = ["", "Material Cost", "2.0x Markup", "2.5x Markup", "3.0x Markup", "3.5x Markup", "4.0x Markup"]
for i, h in enumerate(markup_headers):
    cell = ws3.cell(row=r, column=i+2, value=h)
    cell.font = font_header if i == 0 else font_sub_header
    cell.fill = fill_sub_header
    cell.alignment = align_center
    cell.border = thin_border
r += 1

# Pull from Quote Calculator
step_names = [
    "Step 1: Pre-Emergent + Fert",
    "Step 2: Weed Control + Feed",
    "Step 3: Grub Prevention + Fert",
    "Step 4: Fall Recovery",
    "Step 5: Winterizer",
]
for i, name in enumerate(step_names):
    fill = fill_alt1 if i % 2 == 0 else fill_alt2
    calc_row = 9 + i  # rows 9-13 in Quote Calculator

    ws3.cell(row=r, column=2, value=name).font = font_bold
    ws3.cell(row=r, column=2).fill = fill
    ws3.cell(row=r, column=2).alignment = align_left
    ws3.cell(row=r, column=2).border = thin_border

    # Material cost from calculator
    ws3.cell(row=r, column=3, value=f"='QUOTE CALCULATOR'!E{calc_row}")
    ws3.cell(row=r, column=3).font = font_bold
    ws3.cell(row=r, column=3).fill = fill
    ws3.cell(row=r, column=3).alignment = align_center
    ws3.cell(row=r, column=3).border = thin_border
    ws3.cell(row=r, column=3).number_format = "$#,##0.00"

    # Markup columns
    for j, mult in enumerate([2.0, 2.5, 3.0, 3.5, 4.0]):
        col = 4 + j
        ws3.cell(row=r, column=col, value=f"=C{r}*{mult}")
        ws3.cell(row=r, column=col).font = Font(name="Arial", size=10, bold=True, color=DARK_GREEN)
        ws3.cell(row=r, column=col).fill = fill
        ws3.cell(row=r, column=col).alignment = align_center
        ws3.cell(row=r, column=col).border = thin_border
        ws3.cell(row=r, column=col).number_format = "$#,##0.00"

    r += 1

# Totals row
ws3.cell(row=r, column=2, value="FULL PROGRAM TOTAL").font = Font(name="Arial", size=11, bold=True, color=WHITE)
ws3.cell(row=r, column=2).fill = PatternFill("solid", fgColor=DARK_GREEN)
ws3.cell(row=r, column=2).border = thick_border
for col in range(3, 9):
    ws3.cell(row=r, column=col, value=f"=SUM({get_column_letter(col)}{r-5}:{get_column_letter(col)}{r-1})")
    ws3.cell(row=r, column=col).font = Font(name="Arial", size=12, bold=True, color=WHITE if col == 3 else GOLD)
    ws3.cell(row=r, column=col).fill = PatternFill("solid", fgColor=DARK_GREEN)
    ws3.cell(row=r, column=col).border = thick_border
    ws3.cell(row=r, column=col).number_format = "$#,##0.00"
r += 1

# Profit row
ws3.cell(row=r, column=2, value="PROFIT").font = Font(name="Arial", size=11, bold=True, color="228B22")
ws3.cell(row=r, column=2).fill = fill_profit
ws3.cell(row=r, column=2).border = thick_border
ws3.cell(row=r, column=3).fill = fill_profit
ws3.cell(row=r, column=3).border = thick_border
for col in range(4, 9):
    ws3.cell(row=r, column=col, value=f"={get_column_letter(col)}{r-1}-C{r-1}")
    ws3.cell(row=r, column=col).font = Font(name="Arial", size=12, bold=True, color="228B22")
    ws3.cell(row=r, column=col).fill = fill_profit
    ws3.cell(row=r, column=col).border = thick_border
    ws3.cell(row=r, column=col).number_format = "$#,##0.00"
r += 2

# Section 2: Seasonal revenue projections
merge_and_style(ws3, f"B{r}:H{r}", "SEASONAL REVENUE PROJECTIONS", Font(name="Arial", size=12, bold=True, color=WHITE), fill_header)
r += 1

ws3.merge_cells(f"B{r}:C{r}")
ws3[f"B{r}"].value = "NUMBER OF PROGRAM CUSTOMERS:"
ws3[f"B{r}"].font = font_input_label
ws3[f"B{r}"].alignment = Alignment(horizontal="right", vertical="center")

ws3[f"D{r}"].value = 25
ws3[f"D{r}"].font = font_input
ws3[f"D{r}"].fill = fill_input
ws3[f"D{r}"].border = input_border
ws3[f"D{r}"].alignment = align_center
ws3[f"D{r}"].number_format = "0"
cust_row = r
r += 2

# Revenue table
rev_headers = ["", "Per Customer", "25 Customers", "50 Customers", "75 Customers", "100 Customers", ""]
for i, h in enumerate(rev_headers):
    if i >= 1 and i <= 6:
        cell = ws3.cell(row=r, column=i+1, value=h)
        cell.font = font_sub_header
        cell.fill = fill_sub_header
        cell.alignment = align_center
        cell.border = thin_border
r += 1

# Use the calculator's total for per-customer calcs
rev_rows = [
    ("Total Revenue (2.5x)", 2.5),
    ("Total Revenue (3.0x)", 3.0),
    ("Total Revenue (3.5x)", 3.5),
]
for idx, (label, mult) in enumerate(rev_rows):
    fill = fill_alt1 if idx % 2 == 0 else fill_alt2

    ws3.cell(row=r, column=2, value=label).font = font_bold
    ws3.cell(row=r, column=2).fill = fill
    ws3.cell(row=r, column=2).alignment = align_left
    ws3.cell(row=r, column=2).border = thin_border

    # Per customer = material cost * markup
    ws3.cell(row=r, column=3, value=f"='QUOTE CALCULATOR'!E14*{mult}")
    ws3.cell(row=r, column=3).font = font_bold
    ws3.cell(row=r, column=3).fill = fill
    ws3.cell(row=r, column=3).alignment = align_center
    ws3.cell(row=r, column=3).border = thin_border
    ws3.cell(row=r, column=3).number_format = "$#,##0"

    for j, cust_count in enumerate([25, 50, 75, 100]):
        col = 4 + j
        ws3.cell(row=r, column=col, value=f"=C{r}*{cust_count}")
        ws3.cell(row=r, column=col).font = Font(name="Arial", size=10, bold=True, color=DARK_GREEN)
        ws3.cell(row=r, column=col).fill = fill
        ws3.cell(row=r, column=col).alignment = align_center
        ws3.cell(row=r, column=col).border = thin_border
        ws3.cell(row=r, column=col).number_format = "$#,##0"

    r += 1

# Material cost row
ws3.cell(row=r, column=2, value="Total Material Cost").font = font_bold
ws3.cell(row=r, column=2).fill = fill_cost
ws3.cell(row=r, column=2).alignment = align_left
ws3.cell(row=r, column=2).border = thin_border

ws3.cell(row=r, column=3, value=f"='QUOTE CALCULATOR'!E14")
ws3.cell(row=r, column=3).font = Font(name="Arial", size=10, bold=True, color="CC0000")
ws3.cell(row=r, column=3).fill = fill_cost
ws3.cell(row=r, column=3).alignment = align_center
ws3.cell(row=r, column=3).border = thin_border
ws3.cell(row=r, column=3).number_format = "$#,##0"

for j, cust_count in enumerate([25, 50, 75, 100]):
    col = 4 + j
    ws3.cell(row=r, column=col, value=f"=C{r}*{cust_count}")
    ws3.cell(row=r, column=col).font = Font(name="Arial", size=10, bold=True, color="CC0000")
    ws3.cell(row=r, column=col).fill = fill_cost
    ws3.cell(row=r, column=col).alignment = align_center
    ws3.cell(row=r, column=col).border = thin_border
    ws3.cell(row=r, column=col).number_format = "$#,##0"
r += 2

# Pro tip
merge_and_style(ws3, f"B{r}:H{r}", "PRO TIPS", Font(name="Arial", size=12, bold=True, color=WHITE), fill_header)
r += 1
tips = [
    "The industry standard markup for lawn fertilization is 2.5x–3.0x material cost.",
    "Steps 4 & 5 (fall applications) are what actually build thick lawns. Never skip them.",
    "Upsell aeration + overseeding in September — that's where the real margin is.",
    "Buy product in bulk at start of season. Pallet pricing saves 15-25% at SiteOne.",
    "Grub control (Step 3) can be charged as a separate premium add-on instead of included.",
    "Commercial properties (40K+ sqft) should get volume pricing — lower markup, higher total profit.",
]
for tip in tips:
    merge_and_style(ws3, f"B{r}:H{r}", f"• {tip}", Font(name="Arial", size=9, color="444444"), fill_alt1, align_left)
    r += 1

ws3.freeze_panes = "A5"

# ============================================================
# SAVE
# ============================================================
output_path = r"c:\Users\vance\OneDrive\Desktop\claude-workspace\TotalGuard_Fertilizer_MasterSheet.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
