#!/usr/bin/env python3
"""Generate TotalGuard 2026 Expense Tracker (.xlsx)"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from datetime import date

# === BRANDING (matches Job Log Book) ===
DARK_GREEN = "1B4332"
LIGHT_GREEN = "D8F3DC"
MEDIUM_GREEN = "95D5B2"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
GREEN_YES = "D4EDDA"
RED_NO = "F8D7DA"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=DARK_GREEN)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
title_font = Font(name="Arial", bold=True, color=DARK_GREEN, size=16)
title_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
title_align = Alignment(horizontal="center", vertical="center")
total_font = Font(name="Arial", bold=True, size=12)
total_fill = PatternFill("solid", fgColor=MEDIUM_GREEN)
white_fill = PatternFill("solid", fgColor=WHITE)
gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
)
medium_border = Border(
    left=Side(style="medium"), right=Side(style="medium"),
    top=Side(style="medium"), bottom=Side(style="medium"),
)

CATEGORIES = "Software/SaaS,Equipment,Vehicle/Gas,Vehicle/Maintenance,Insurance,Licensing/Permits,Marketing/Advertising,Office Supplies,Phone/Communications,Professional Services,Subcontractors,Training/Education,Travel,Meals (50%),Repairs/Maintenance,Uniforms/Safety,Other"
PAYMENT_METHODS = "Cash,Check,Venmo,Zelle,Credit Card,Debit Card,ACH/Bank"
RECURRING = "One-Time,Monthly,Annual"
YES_NO = "Yes,No"
RECEIPT = "Yes,No,N/A"
SCHEDULE_C = "Line 8 - Advertising,Line 10 - Car/Truck Expenses,Line 15 - Insurance,Line 17 - Legal/Professional,Line 18 - Office Expense,Line 20a - Rent (Vehicles/Equipment),Line 22 - Supplies,Line 24a - Travel,Line 24b - Meals,Line 25 - Utilities,Line 27a - Other Expenses"
VEHICLES = "Car,Truck,Trailer"
MILEAGE_PURPOSE = "Client Visit,Estimate,Supply Run,Job Site,Equipment Pickup,Bank/Admin,Other"

# Pre-populated expense data (matched exactly to current Google Sheet)
EXPENSES_DATA = [
    (date(2026, 1, 10), "Phone/Communications", "Quo (phone app)", 19.59, "Debit Card", "Monthly", "Yes", "Line 25 - Utilities"),
    (date(2026, 1, 10), "Phone/Communications", "Quo (phone app)", 19.50, "Debit Card", "Monthly", "Yes", "Line 25 - Utilities"),
    (date(2026, 1, 14), "Phone/Communications", "Quo", 1.00, "Debit Card", "Monthly", "Yes", "Line 25 - Utilities"),
    (date(2026, 1, 29), "Software/SaaS", "Lovable Website Builder", 15.00, "Debit Card", "Monthly", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 1, 30), "Software/SaaS", "Chat subscription", 20.00, "Debit Card", "Monthly", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 1, 31), "Software/SaaS", "Lovable Website Builder", 12.50, "Debit Card", "Monthly", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 2, 4), "Software/SaaS", "Brevo (marketing)", 9.00, "Debit Card", "Monthly", "Yes", "Line 8 - Advertising"),
    (date(2026, 2, 6), "Marketing/Advertising", "Reviews ($5/per)", 45.00, "Venmo", "One-Time", "Yes", "Line 8 - Advertising"),
    (date(2026, 2, 9), "Phone/Communications", "Quo (Biz phone)", 22.01, "ACH/Bank", "Monthly", "Yes", "Line 25 - Utilities"),
    (date(2026, 2, 9), "Phone/Communications", "Quo", 22.01, "ACH/Bank", "Monthly", "Yes", "Line 25 - Utilities"),
    (date(2026, 2, 12), "Software/SaaS", "Lovable Website Builder", 30.00, "Debit Card", "Monthly", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 2, 23), "Software/SaaS", "Claude Code", 100.00, "Debit Card", "Monthly", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 2, 24), "Training/Education", "Pearson VUE exam", 45.00, "Debit Card", "One-Time", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 2, 28), "Software/SaaS", "Lovable Website Builder", 13.00, "Debit Card", "Monthly", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 2, 28), "Equipment", "Monitor", 105.00, "Debit Card", "One-Time", "Yes", "Line 22 - Supplies"),
    (date(2026, 3, 5), "Software/SaaS", "Jobber Subscription", 49.00, "Debit Card", "Monthly", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 3, 5), "Licensing/Permits", "Licence application fee", 45.00, "Debit Card", "One-Time", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 3, 7), "Insurance", "Pekin Liability Insurance", 86.00, "ACH/Bank", "Annual", "Yes", "Line 15 - Insurance"),
    (date(2026, 3, 7), "Software/SaaS", "Porkbun Domains", 22.00, "Debit Card", "Annual", "Yes", "Line 27a - Other Expenses"),
    (date(2026, 3, 8), "Marketing/Advertising", "1x G Review", 5.00, "Venmo", "One-Time", "Yes", "Line 8 - Advertising"),
    (date(2026, 3, 9), "Phone/Communications", "Twilio", 20.00, "Debit Card", "Monthly", "Yes", "Line 25 - Utilities"),
    (date(2026, 3, 9), "Phone/Communications", "Quo", 22.01, "ACH/Bank", "Monthly", "Yes", "Line 25 - Utilities"),
    (date(2026, 3, 10), "Marketing/Advertising", "Yard Signs Uz Marketing", 398.00, "Check", "One-Time", "Yes", "Line 8 - Advertising"),
    (date(2026, 3, 11), "Equipment", "Trailer 5 x 10 (dad)", 1582.00, "Check", "One-Time", "Yes", "Line 22 - Supplies"),
    (date(2026, 3, 11), "Equipment", "Trailer Hooks", 40.00, "Cash", "One-Time", "Yes", "Line 22 - Supplies"),
    (date(2026, 3, 11), "Software/SaaS", "n8n", 24.00, "Debit Card", "Monthly", "Yes", "Line 27a - Other Expenses"),
]

GAS_DATA = [
    (date(2026, 1, 1), "Car", 22.68),
    (date(2026, 1, 8), "Car", 25.82),
    (date(2026, 1, 12), "Car", 38.04),
    (date(2026, 1, 27), "Car", 30.87),
    (date(2026, 2, 2), "Car", 30.68),
    (date(2026, 2, 6), "Car", 32.13),
    (date(2026, 3, 3), "Car", 35.23),
    (date(2026, 3, 11), "Car", 44.25),
]


def style_data_rows(ws, start_row, end_row, num_cols):
    for r in range(start_row, end_row + 1):
        fill = white_fill if (r % 2 == 1) else gray_fill
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")


def add_title(ws, title_text, merge_range):
    ws.merge_cells(merge_range)
    end_col = merge_range.split(":")[1][0]
    cell = ws[merge_range.split(":")[0]]
    cell.value = title_text
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = title_align
    ws.row_dimensions[1].height = 36


def add_headers(ws, headers, widths):
    ws.row_dimensions[2].height = 28
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width


def add_dv(ws, formula, col_letter, start_row, end_row, prompt_title="", prompt=""):
    dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    if prompt_title:
        dv.promptTitle = prompt_title
        dv.prompt = prompt
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def create_expenses_tab(wb):
    ws = wb.create_sheet(title="Expenses")
    add_title(ws, "TOTALGUARD YARD CARE \u2014 2026 EXPENSE TRACKER", "A1:K1")

    headers = ["#", "Date", "Category", "Vendor/Description", "Amount", "Payment Method",
               "Recurring?", "Tax Deductible?", "Schedule C Category", "Receipt?", "Notes"]
    widths = [6, 14, 20, 30, 14, 16, 14, 16, 24, 10, 28]
    add_headers(ws, headers, widths)

    DATA_END = 502
    style_data_rows(ws, 3, DATA_END, 11)

    # Pre-fill row numbers
    for i in range(500):
        ws.cell(row=3 + i, column=1, value=i + 1).alignment = Alignment(horizontal="center", vertical="center")

    # Date and currency formatting for all rows
    for r in range(3, DATA_END + 1):
        ws.cell(row=r, column=2).number_format = "MM/DD/YYYY"
        ws.cell(row=r, column=5).number_format = '$#,##0.00'
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center")

    # Pre-populate data
    for i, (dt, cat, vendor, amt, pay, recur, deduct, sched_c) in enumerate(EXPENSES_DATA):
        row = 3 + i
        ws.cell(row=row, column=2, value=dt)
        ws.cell(row=row, column=3, value=cat)
        ws.cell(row=row, column=4, value=vendor)
        ws.cell(row=row, column=5, value=amt)
        ws.cell(row=row, column=6, value=pay)
        ws.cell(row=row, column=7, value=recur)
        ws.cell(row=row, column=8, value=deduct)
        ws.cell(row=row, column=9, value=sched_c)
        ws.cell(row=row, column=10, value="No")

    # Total row
    total_row = 503
    ws.cell(row=total_row, column=4, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="right", vertical="center")
    total_cell = ws.cell(row=total_row, column=5)
    total_cell.value = f"=SUM(E3:E{DATA_END})"
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.number_format = '$#,##0.00'
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.border = medium_border

    # Data validations
    add_dv(ws, CATEGORIES, "C", 3, DATA_END, "Category", "Select expense category")
    add_dv(ws, PAYMENT_METHODS, "F", 3, DATA_END, "Payment", "Select payment method")
    add_dv(ws, RECURRING, "G", 3, DATA_END, "Recurring", "One-Time, Monthly, or Annual")
    add_dv(ws, YES_NO, "H", 3, DATA_END, "Tax Deductible", "Is this tax deductible?")
    add_dv(ws, SCHEDULE_C, "I", 3, DATA_END, "Schedule C", "IRS Schedule C line item")
    add_dv(ws, RECEIPT, "J", 3, DATA_END, "Receipt", "Do you have a receipt?")

    # Conditional formatting on Tax Deductible column
    ws.conditional_formatting.add(
        f"H3:H{DATA_END}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor=GREEN_YES))
    )
    ws.conditional_formatting.add(
        f"H3:H{DATA_END}",
        CellIsRule(operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor=RED_NO))
    )

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"
    return ws


def create_gas_tab(wb):
    ws = wb.create_sheet(title="Gas & Fuel")
    add_title(ws, "TOTALGUARD YARD CARE \u2014 2026 GAS & FUEL LOG", "A1:H1")

    headers = ["#", "Date", "Vehicle", "Gallons", "Price/Gal", "Total Cost", "Odometer", "Notes"]
    widths = [6, 14, 12, 10, 12, 14, 12, 24]
    add_headers(ws, headers, widths)

    DATA_END = 202
    style_data_rows(ws, 3, DATA_END, 8)

    for i in range(200):
        r = 3 + i
        ws.cell(row=r, column=1, value=i + 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).number_format = "MM/DD/YYYY"
        ws.cell(row=r, column=4).number_format = '#,##0.000'
        ws.cell(row=r, column=5).number_format = '$#,##0.000'
        ws.cell(row=r, column=6).number_format = '$#,##0.00'
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r, column=7).number_format = '#,##0'

    # Pre-populate gas data
    for i, (dt, vehicle, cost) in enumerate(GAS_DATA):
        row = 3 + i
        ws.cell(row=row, column=2, value=dt)
        ws.cell(row=row, column=3, value=vehicle)
        ws.cell(row=row, column=6, value=cost)

    # Total row
    total_row = 203
    ws.cell(row=total_row, column=5, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    total_cell = ws.cell(row=total_row, column=6)
    total_cell.value = f"=SUM(F3:F{DATA_END})"
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.number_format = '$#,##0.00'
    total_cell.border = medium_border

    add_dv(ws, VEHICLES, "C", 3, DATA_END, "Vehicle", "Select vehicle")

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"
    return ws


def create_mileage_tab(wb):
    ws = wb.create_sheet(title="Mileage")
    add_title(ws, "TOTALGUARD YARD CARE \u2014 2026 MILEAGE LOG", "A1:I1")

    headers = ["#", "Date", "Purpose", "From", "To", "Start Miles", "End Miles", "Miles Driven", "Deduction"]
    widths = [6, 14, 18, 22, 22, 12, 12, 14, 14]
    add_headers(ws, headers, widths)

    DATA_END = 302
    style_data_rows(ws, 3, DATA_END, 9)

    for i in range(300):
        r = 3 + i
        ws.cell(row=r, column=1, value=i + 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).number_format = "MM/DD/YYYY"
        ws.cell(row=r, column=6).number_format = '#,##0.0'
        ws.cell(row=r, column=7).number_format = '#,##0.0'
        ws.cell(row=r, column=8).number_format = '#,##0.1'
        ws.cell(row=r, column=8).value = f'=IF(AND(F{r}<>"",G{r}<>""),G{r}-F{r},"")'
        ws.cell(row=r, column=9).number_format = '$#,##0.00'
        ws.cell(row=r, column=9).value = f'=IF(H{r}<>"",H{r}*0.70,"")'
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="right", vertical="center")

    # Total row
    total_row = 303
    ws.cell(row=total_row, column=7, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=7).alignment = Alignment(horizontal="right", vertical="center")

    miles_total = ws.cell(row=total_row, column=8)
    miles_total.value = f"=SUM(H3:H{DATA_END})"
    miles_total.font = total_font
    miles_total.fill = total_fill
    miles_total.number_format = '#,##0.1'
    miles_total.border = medium_border

    deduction_total = ws.cell(row=total_row, column=9)
    deduction_total.value = f"=SUM(I3:I{DATA_END})"
    deduction_total.font = total_font
    deduction_total.fill = total_fill
    deduction_total.number_format = '$#,##0.00'
    deduction_total.border = medium_border

    # IRS note
    note_cell = ws.cell(row=305, column=6)
    note_cell.value = "IRS Standard Mileage Rate 2026: $0.70/mile"
    note_cell.font = Font(name="Arial", italic=True, color=DARK_GREEN, size=10)
    ws.merge_cells("F305:I305")

    add_dv(ws, MILEAGE_PURPOSE, "C", 3, DATA_END, "Purpose", "Trip purpose")

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"
    return ws


def create_monthly_summary(wb):
    ws = wb.create_sheet(title="Monthly Summary")
    add_title(ws, "TOTALGUARD YARD CARE \u2014 2026 MONTHLY EXPENSE SUMMARY", "A1:G1")

    headers = ["Month", "General Expenses", "Gas & Fuel", "Mileage Deduction", "Total Spent", "Total Deductions", "YTD Total"]
    widths = [14, 18, 14, 18, 14, 18, 14]
    add_headers(ws, headers, widths)

    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]

    style_data_rows(ws, 3, 14, 7)

    for i, month in enumerate(months):
        row = 3 + i
        month_num = i + 1
        ws.cell(row=row, column=1, value=month).font = Font(name="Arial", bold=True, size=11)

        # General Expenses: SUMPRODUCT matching month from Expenses tab
        ws.cell(row=row, column=2).value = f'=SUMPRODUCT((MONTH(Expenses!B$3:B$502)={month_num})*(Expenses!E$3:E$502))'
        ws.cell(row=row, column=2).number_format = '$#,##0.00'

        # Gas & Fuel: SUMPRODUCT matching month from Gas & Fuel tab
        ws.cell(row=row, column=3).value = f"=SUMPRODUCT((MONTH('Gas & Fuel'!B$3:B$202)={month_num})*('Gas & Fuel'!F$3:F$202))"
        ws.cell(row=row, column=3).number_format = '$#,##0.00'

        # Mileage Deduction: SUMPRODUCT matching month from Mileage tab
        ws.cell(row=row, column=4).value = f'=SUMPRODUCT((MONTH(Mileage!B$3:B$302)={month_num})*(Mileage!I$3:I$302))'
        ws.cell(row=row, column=4).number_format = '$#,##0.00'

        # Total Spent = General + Gas
        ws.cell(row=row, column=5).value = f'=B{row}+C{row}'
        ws.cell(row=row, column=5).number_format = '$#,##0.00'

        # Total Deductions = General + Gas + Mileage
        ws.cell(row=row, column=6).value = f'=B{row}+C{row}+D{row}'
        ws.cell(row=row, column=6).number_format = '$#,##0.00'

        # YTD Total = running sum of Total Spent
        ws.cell(row=row, column=7).value = f'=SUM(E$3:E{row})'
        ws.cell(row=row, column=7).number_format = '$#,##0.00'

    # Annual total row
    total_row = 15
    for c in range(1, 8):
        ws.cell(row=total_row, column=c).fill = total_fill
        ws.cell(row=total_row, column=c).border = medium_border
        ws.cell(row=total_row, column=c).font = total_font
    ws.cell(row=total_row, column=1, value="ANNUAL TOTAL")
    for c in range(2, 8):
        ws.cell(row=total_row, column=c).value = f'=SUM({get_column_letter(c)}3:{get_column_letter(c)}14)'
        ws.cell(row=total_row, column=c).number_format = '$#,##0.00'

    # Tax savings estimate
    ws.cell(row=17, column=5, value="Estimated Tax Savings (25%)").font = Font(name="Arial", bold=True, italic=True, size=11, color=DARK_GREEN)
    ws.cell(row=17, column=6).value = '=F15*0.25'
    ws.cell(row=17, column=6).number_format = '$#,##0.00'
    ws.cell(row=17, column=6).font = Font(name="Arial", bold=True, size=14, color=DARK_GREEN)
    ws.cell(row=17, column=6).border = medium_border

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:2"
    return ws


def create_tax_summary(wb):
    ws = wb.create_sheet(title="Tax Summary")
    add_title(ws, "TOTALGUARD YARD CARE \u2014 2026 SCHEDULE C TAX SUMMARY", "A1:C1")

    headers = ["Schedule C Line", "Total Amount", "% of Total"]
    widths = [30, 16, 12]
    add_headers(ws, headers, widths)

    schedule_c_items = [
        "Line 8 - Advertising",
        "Line 10 - Car/Truck Expenses",
        "Line 15 - Insurance",
        "Line 17 - Legal/Professional",
        "Line 18 - Office Expense",
        "Line 20a - Rent (Vehicles/Equipment)",
        "Line 22 - Supplies",
        "Line 24a - Travel",
        "Line 24b - Meals",
        "Line 25 - Utilities",
        "Line 27a - Other Expenses",
    ]

    style_data_rows(ws, 3, 16, 3)

    for i, item in enumerate(schedule_c_items):
        row = 3 + i
        ws.cell(row=row, column=1, value=item).font = Font(name="Arial", size=11)
        ws.cell(row=row, column=2).value = f'=SUMPRODUCT((Expenses!I$3:I$502="{item}")*(Expenses!E$3:E$502))'
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=3).number_format = '0.0%'

    # Vehicle Mileage Deduction (row 14)
    ws.cell(row=14, column=1, value="Vehicle Mileage Deduction").font = Font(name="Arial", size=11)
    ws.cell(row=14, column=2).value = "=Mileage!I303"
    ws.cell(row=14, column=2).number_format = '$#,##0.00'
    ws.cell(row=14, column=3).number_format = '0.0%'

    # Gas & Fuel (row 15)
    ws.cell(row=15, column=1, value="Gas & Fuel").font = Font(name="Arial", size=11)
    ws.cell(row=15, column=2).value = "='Gas & Fuel'!F203"
    ws.cell(row=15, column=2).number_format = '$#,##0.00'
    ws.cell(row=15, column=3).number_format = '0.0%'

    # Total deductions row
    total_row = 17
    for c in range(1, 4):
        ws.cell(row=total_row, column=c).fill = total_fill
        ws.cell(row=total_row, column=c).border = medium_border
        ws.cell(row=total_row, column=c).font = total_font
    ws.cell(row=total_row, column=1, value="TOTAL DEDUCTIONS")
    ws.cell(row=total_row, column=2).value = "=SUM(B3:B15)"
    ws.cell(row=total_row, column=2).number_format = '$#,##0.00'

    # Percentage formulas (reference total)
    for r in range(3, 16):
        ws.cell(row=r, column=3).value = f'=IF(B$17>0,B{r}/B$17,0)'

    ws.cell(row=total_row, column=3).value = "=SUM(C3:C15)"
    ws.cell(row=total_row, column=3).number_format = '0.0%'

    # Disclaimer
    ws.cell(row=19, column=1, value="Note: Consult a tax professional. This is for tracking purposes only.").font = Font(name="Arial", italic=True, size=9, color="888888")
    ws.merge_cells("A19:C19")

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:2"
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    create_expenses_tab(wb)
    create_gas_tab(wb)
    create_mileage_tab(wb)
    create_monthly_summary(wb)
    create_tax_summary(wb)

    output_path = "TotalGuard_2026_Expense_Tracker.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    main()
