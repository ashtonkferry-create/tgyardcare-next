#!/usr/bin/env python3
"""Generate TotalGuard 2026 Job Log Book (.xlsx)"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import calendar

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

SERVICES = "Lawn Mowing,Fertilization,Aeration,Herbicide,Weeding,Mulching,Garden Bed Care,Bush Trimming,Hardscaping,Spring Cleanup,Fall Cleanup,Leaf Removal,Gutter Cleaning,Gutter Guard Installation,Snow Removal"
CITIES = "Madison,Middleton,Waunakee,Monona,Sun Prairie,Fitchburg,Verona,McFarland,Cottage Grove,DeForest,Oregon,Stoughton"
CREW = "Vance,Partner,Both"
PAYMENT = "Cash,Check,Venmo,Zelle,Credit Card,Invoice"
PAID = "YES,NO"

HEADERS = ["Job #", "Date", "Crew", "Customer", "Service", "Address", "City", "Contact", "Payment Method", "Amount", "Paid?", "Notes"]
WIDTHS = [8, 14, 12, 24, 22, 32, 16, 22, 18, 14, 10, 30]

DARK_GREEN = "1B4332"
LIGHT_GREEN = "D8F3DC"
MEDIUM_GREEN = "95D5B2"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
GREEN_PAID = "D4EDDA"
RED_UNPAID = "F8D7DA"

DATA_ROWS = 100  # rows 3-102

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=DARK_GREEN)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

title_font = Font(name="Arial", bold=True, color=DARK_GREEN, size=16)
title_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
title_align = Alignment(horizontal="center", vertical="center")

total_font = Font(name="Arial", bold=True, size=12)
total_fill = PatternFill("solid", fgColor=MEDIUM_GREEN)

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

white_fill = PatternFill("solid", fgColor=WHITE)
gray_fill = PatternFill("solid", fgColor=LIGHT_GRAY)

green_fill = PatternFill("solid", fgColor=GREEN_PAID)
red_fill = PatternFill("solid", fgColor=RED_UNPAID)


def create_monthly_sheet(wb, month_name):
    ws = wb.create_sheet(title=f"{month_name} 2026")

    # Column widths
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 24

    # Title row (merged A1:L1)
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = f"TOTALGUARD YARD CARE \u2014 {month_name.upper()} 2026"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_align

    # Header row
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows (3-102): alternating colors, borders, job numbers
    for r in range(3, 3 + DATA_ROWS):
        fill = white_fill if (r % 2 == 1) else gray_fill
        for c in range(1, 13):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

        # Job # column: center-aligned
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
        # Date column: date format
        ws.cell(row=r, column=2).number_format = "MM/DD/YYYY"
        # Amount column: currency
        ws.cell(row=r, column=10).number_format = '$#,##0.00'
        ws.cell(row=r, column=10).alignment = Alignment(horizontal="right", vertical="center")

    # Pre-fill job numbers 1-100
    for i in range(DATA_ROWS):
        ws.cell(row=3 + i, column=1, value=i + 1)

    # Total row (row 103)
    total_row = 3 + DATA_ROWS
    ws.cell(row=total_row, column=9, value="TOTAL").font = total_font
    ws.cell(row=total_row, column=9).alignment = Alignment(horizontal="right", vertical="center")

    total_cell = ws.cell(row=total_row, column=10)
    total_cell.value = f"=SUM(J3:J{total_row - 1})"
    total_cell.font = total_font
    total_cell.fill = total_fill
    total_cell.number_format = '$#,##0.00'
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.border = Border(
        left=Side(style="medium"), right=Side(style="medium"),
        top=Side(style="medium"), bottom=Side(style="medium"),
    )

    # Data validations
    dv_crew = DataValidation(type="list", formula1=f'"{CREW}"', allow_blank=True)
    dv_crew.prompt = "Select crew member"
    dv_crew.promptTitle = "Crew"
    ws.add_data_validation(dv_crew)
    dv_crew.add(f"C3:C{total_row - 1}")

    dv_service = DataValidation(type="list", formula1=f'"{SERVICES}"', allow_blank=True)
    dv_service.prompt = "Select service type"
    dv_service.promptTitle = "Service"
    ws.add_data_validation(dv_service)
    dv_service.add(f"E3:E{total_row - 1}")

    dv_city = DataValidation(type="list", formula1=f'"{CITIES}"', allow_blank=True)
    dv_city.prompt = "Select service area"
    dv_city.promptTitle = "City"
    ws.add_data_validation(dv_city)
    dv_city.add(f"G3:G{total_row - 1}")

    dv_payment = DataValidation(type="list", formula1=f'"{PAYMENT}"', allow_blank=True)
    dv_payment.prompt = "Select payment method"
    dv_payment.promptTitle = "Payment"
    ws.add_data_validation(dv_payment)
    dv_payment.add(f"I3:I{total_row - 1}")

    dv_paid = DataValidation(type="list", formula1=f'"{PAID}"', allow_blank=True)
    dv_paid.prompt = "Paid?"
    dv_paid.promptTitle = "Payment Status"
    ws.add_data_validation(dv_paid)
    dv_paid.add(f"K3:K{total_row - 1}")

    # Conditional formatting for Paid? column
    ws.conditional_formatting.add(
        f"K3:K{total_row - 1}",
        CellIsRule(operator="equal", formula=['"YES"'], fill=PatternFill("solid", fgColor=GREEN_PAID))
    )
    ws.conditional_formatting.add(
        f"K3:K{total_row - 1}",
        CellIsRule(operator="equal", formula=['"NO"'], fill=PatternFill("solid", fgColor=RED_UNPAID))
    )

    # Freeze panes (row 3, so rows 1-2 stay visible)
    ws.freeze_panes = "A3"

    # Print setup
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"

    return ws


def create_summary_sheet(wb):
    ws = wb.create_sheet(title="Summary")

    # Column widths
    summary_widths = [22, 16, 18, 12, 12]
    for i, w in enumerate(summary_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "TOTALGUARD YARD CARE \u2014 2026 ANNUAL SUMMARY"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_align
    ws.row_dimensions[1].height = 36

    # Headers
    summary_headers = ["Month", "Total Jobs", "Total Revenue", "Paid", "Unpaid"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Monthly rows
    for i, month in enumerate(MONTHS):
        row = 3 + i
        sheet_name = f"'{month} 2026'"
        fill = white_fill if (row % 2 == 1) else gray_fill

        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).alignment = Alignment(vertical="center")

        ws.cell(row=row, column=1, value=month).font = Font(name="Arial", bold=True, size=11)
        ws.cell(row=row, column=2).value = f"=COUNTA({sheet_name}!D3:D102)"
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=3).value = f"={sheet_name}!J103"
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        ws.cell(row=row, column=4).value = f'=COUNTIF({sheet_name}!K3:K102,"YES")'
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=5).value = f'=COUNTIF({sheet_name}!K3:K102,"NO")'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")

    # Grand total row
    total_row = 15
    for c in range(1, 6):
        ws.cell(row=total_row, column=c).fill = total_fill
        ws.cell(row=total_row, column=c).border = Border(
            left=Side(style="medium"), right=Side(style="medium"),
            top=Side(style="medium"), bottom=Side(style="medium"),
        )
        ws.cell(row=total_row, column=c).font = total_font

    ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    ws.cell(row=total_row, column=2).value = "=SUM(B3:B14)"
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=3).value = "=SUM(C3:C14)"
    ws.cell(row=total_row, column=3).number_format = '$#,##0.00'
    ws.cell(row=total_row, column=4).value = "=SUM(D3:D14)"
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=5).value = "=SUM(E3:E14)"
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A3"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    return ws


def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for month in MONTHS:
        create_monthly_sheet(wb, month)

    create_summary_sheet(wb)

    output_path = "TotalGuard_2026_Job_Log_Book.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")


if __name__ == "__main__":
    main()
