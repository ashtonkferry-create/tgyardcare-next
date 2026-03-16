#!/usr/bin/env python3
"""
Build TotalGuard 2026 Equipment & Asset Inventory spreadsheet.
Generates TotalGuard_2026_Equipment_Inventory.xlsx with two tabs:
  1. Equipment & Assets — 100-row capacity (11 columns, no depreciation)
  2. Maintenance Log — 200-row capacity
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ── Branding constants ──────────────────────────────────────────────
DARK_GREEN = "1B4332"
LIGHT_GREEN = "D8F3DC"
MEDIUM_GREEN = "95D5B2"
WHITE = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
BORDER_COLOR = "CCCCCC"

HEADER_FILL = PatternFill(start_color=DARK_GREEN, end_color=DARK_GREEN, fill_type="solid")
TITLE_FILL = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
TOTAL_FILL = PatternFill(start_color=MEDIUM_GREEN, end_color=MEDIUM_GREEN, fill_type="solid")
WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
GRAY_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name="Arial", bold=True, color=DARK_GREEN, size=16)
DATA_FONT = Font(name="Arial", size=11)
TOTAL_FONT = Font(name="Arial", bold=True, size=11)

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
CURRENCY_FMT = '$#,##0.00'
DATE_FMT = 'MM/DD/YYYY'

# ── Conditional formatting fills ────────────────────────────────────
COND_GREEN = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
COND_YELLOW = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
COND_RED = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
COND_GRAY = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")


def apply_row_style(ws, row, col_count, is_alt, font=None, fill_override=None):
    """Apply font, fill, and border to a row."""
    f = font or DATA_FONT
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = f
        cell.border = THIN_BORDER
        if fill_override:
            cell.fill = fill_override
        elif is_alt:
            cell.fill = GRAY_FILL
        else:
            cell.fill = WHITE_FILL


def build_equipment_tab(wb):
    ws = wb.active
    ws.title = "Equipment & Assets"

    col_count = 11  # A-K

    # ── Column widths ───────────────────────────────────────────────
    widths = [8, 28, 16, 22, 18, 14, 14, 12, 14, 12, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row 1: Title ────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1,
                         value="TOTALGUARD YARD CARE — 2026 EQUIPMENT & ASSET INVENTORY")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = TITLE_FILL
        cell.border = THIN_BORDER

    # ── Row 2: Headers ──────────────────────────────────────────────
    headers = [
        "Asset #", "Item Name", "Category", "Brand/Model", "Serial #",
        "Purchase Date", "Purchase Price", "Condition", "Warranty Expiry",
        "Location", "Notes"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ── Data rows 3-102 (100 capacity) ──────────────────────────────
    # Dropdowns
    cat_dv = DataValidation(
        type="list",
        formula1='"Mower,Trimmer,Blower,Edger,Trailer,Vehicle,Hand Tools,Safety Gear,Technology,Office Equipment,Sprayer,Aerator,Other"',
        allow_blank=True
    )
    cond_dv = DataValidation(
        type="list",
        formula1='"New,Excellent,Good,Fair,Poor,Retired"',
        allow_blank=True
    )
    loc_dv = DataValidation(
        type="list",
        formula1='"Truck,Garage,Job Site,Storage,Office"',
        allow_blank=True
    )
    ws.add_data_validation(cat_dv)
    ws.add_data_validation(cond_dv)
    ws.add_data_validation(loc_dv)

    for row in range(3, 103):
        is_alt = (row - 3) % 2 == 1
        # Col A: Asset #
        ws.cell(row=row, column=1, value=row - 2).alignment = CENTER
        # Col F: date format
        ws.cell(row=row, column=6).number_format = DATE_FMT
        # Col G: currency
        ws.cell(row=row, column=7).number_format = CURRENCY_FMT
        # Col I: date format
        ws.cell(row=row, column=9).number_format = DATE_FMT

        apply_row_style(ws, row, col_count, is_alt)

        # Dropdowns
        cat_dv.add(ws.cell(row=row, column=3))
        cond_dv.add(ws.cell(row=row, column=8))
        loc_dv.add(ws.cell(row=row, column=10))

    # ── Pre-populate known equipment ────────────────────────────────
    equipment = [
        # (row, name, category, brand, serial, date, price, condition, warranty, location, notes)
        (3, "Trailer 5x10", "Trailer", "", "", "03/11/2026", 1582.00, "Good", "", "Truck", ""),
        (4, "Monitor", "Technology", "", "", "02/28/2026", 105.00, "New", "", "Office", ""),
        (5, "Trailer Hooks", "Hand Tools", "", "", "03/11/2026", 40.00, "New", "", "Truck", ""),
        (6, "Yard Signs", "Other", "", "", "03/10/2026", 398.00, "New", "", "", "Marketing materials"),
    ]
    for item in equipment:
        row = item[0]
        ws.cell(row=row, column=2, value=item[1])   # Item Name
        ws.cell(row=row, column=3, value=item[2])   # Category
        ws.cell(row=row, column=4, value=item[3])   # Brand/Model
        ws.cell(row=row, column=5, value=item[4])   # Serial #
        ws.cell(row=row, column=6, value=item[5])   # Purchase Date
        ws.cell(row=row, column=6).number_format = DATE_FMT
        ws.cell(row=row, column=7, value=item[6])   # Purchase Price
        ws.cell(row=row, column=7).number_format = CURRENCY_FMT
        ws.cell(row=row, column=8, value=item[7])   # Condition
        ws.cell(row=row, column=9, value=item[8])   # Warranty Expiry
        ws.cell(row=row, column=10, value=item[9])  # Location
        ws.cell(row=row, column=11, value=item[10]) # Notes

    # ── Row 103: TOTAL row ──────────────────────────────────────────
    total_row = 103
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).alignment = CENTER
    ws.cell(row=total_row, column=7, value="=SUM(G3:G102)")
    ws.cell(row=total_row, column=7).number_format = CURRENCY_FMT
    apply_row_style(ws, total_row, col_count, False, font=TOTAL_FONT, fill_override=TOTAL_FILL)

    # ── Conditional formatting on column H (Condition) ──────────────
    h_range = "H3:H102"
    ws.conditional_formatting.add(h_range, CellIsRule(
        operator="equal", formula=['"New"'], fill=COND_GREEN))
    ws.conditional_formatting.add(h_range, CellIsRule(
        operator="equal", formula=['"Excellent"'], fill=COND_GREEN))
    ws.conditional_formatting.add(h_range, CellIsRule(
        operator="equal", formula=['"Fair"'], fill=COND_YELLOW))
    ws.conditional_formatting.add(h_range, CellIsRule(
        operator="equal", formula=['"Poor"'], fill=COND_RED))
    ws.conditional_formatting.add(h_range, CellIsRule(
        operator="equal", formula=['"Retired"'], fill=COND_GRAY))

    # ── Freeze panes below headers ──────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Print setup ─────────────────────────────────────────────────
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"


def build_maintenance_tab(wb):
    ws = wb.create_sheet("Maintenance Log")

    col_count = 8  # A-H

    # ── Column widths ───────────────────────────────────────────────
    widths = [6, 14, 10, 24, 18, 12, 14, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row 1: Title ────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1,
                         value="TOTALGUARD YARD CARE — 2026 MAINTENANCE LOG")
    title_cell.font = TITLE_FONT
    title_cell.fill = TITLE_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = TITLE_FILL
        cell.border = THIN_BORDER

    # ── Row 2: Headers ──────────────────────────────────────────────
    headers = [
        "#", "Date", "Asset #", "Item Name", "Maintenance Type",
        "Cost", "Performed By", "Notes"
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ── Dropdowns ───────────────────────────────────────────────────
    maint_dv = DataValidation(
        type="list",
        formula1='"Oil Change,Blade Sharpening,Belt Replace,Filter Change,Tire Repair,General Service,Repair,Inspection,Winterize,Other"',
        allow_blank=True
    )
    perf_dv = DataValidation(
        type="list",
        formula1='"Self,Dealer,Mechanic,Other"',
        allow_blank=True
    )
    ws.add_data_validation(maint_dv)
    ws.add_data_validation(perf_dv)

    # ── Data rows 3-202 (200 capacity) ──────────────────────────────
    for row in range(3, 203):
        is_alt = (row - 3) % 2 == 1
        # Col A: #
        ws.cell(row=row, column=1, value=row - 2).alignment = CENTER
        # Col B: date
        ws.cell(row=row, column=2).number_format = DATE_FMT
        # Col C: Asset # (center)
        ws.cell(row=row, column=3).alignment = CENTER
        # Col F: currency
        ws.cell(row=row, column=6).number_format = CURRENCY_FMT

        apply_row_style(ws, row, col_count, is_alt)

        maint_dv.add(ws.cell(row=row, column=5))
        perf_dv.add(ws.cell(row=row, column=7))

    # ── Row 203: TOTAL row ──────────────────────────────────────────
    total_row = 203
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).alignment = CENTER
    ws.cell(row=total_row, column=6, value="=SUM(F3:F202)")
    ws.cell(row=total_row, column=6).number_format = CURRENCY_FMT
    apply_row_style(ws, total_row, col_count, False, font=TOTAL_FONT, fill_override=TOTAL_FILL)

    # ── Freeze panes below headers ──────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Print setup ─────────────────────────────────────────────────
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:2"


def main():
    wb = Workbook()
    build_equipment_tab(wb)
    build_maintenance_tab(wb)

    out_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path = os.path.join(out_dir, "TotalGuard_2026_Equipment_Inventory.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
